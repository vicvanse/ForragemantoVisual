"""
Forrageamento visual Exp1 — PsychoPy Coder.

Compatível com PsychoPy 2021.2.x (Python 3.6) e 2023.x (Python 3.8+).

Requer: EyeLink SDK + pylink no PYTHONPATH para modo EyeLink real.
Coloque EyeLinkCoreGraphicsPsychoPy.py na mesma pasta (sem alterações).

Uso:
  python forrageamento_exp1.py                    # pede participante, Experimento 1 ou 2, sessão, repetição; menu E/M (se aplicável)
  python forrageamento_exp1.py --participant P01 --experiment 2 --session 3 --session-run 1
  # FORR_PARTICIPANT_ID, FORR_EXPERIMENT (1 ou 2), FORR_SESSION (1–5), FORR_SESSION_RUN (>=1) — idem
  python forrageamento_exp1.py --dummy           # força rato, sem menu
  python forrageamento_exp1.py --skip-mode-menu  # sem menu: só EyeLink (ou --dummy)
  python forrageamento_exp1.py --trials 5

Janela (por defeito 1680×1050): --window-width / --window-height, --resolution WxH,
  FORR_WIN_W / FORR_WIN_H ou FORR_RESOLUTION (ex.: 1680x1050). No diálogo gráfico pode escolher-se o preset.
Estímulos são gerados no código (L e T dispersos no painel, tons de cinza). Instruções: texto branco em PT-BR; exemplo procedural (caixa com L e T verde).

Metadados no início (diálogo ou --participant / --experiment / --session / --session-run / FORR_*):
  (1) participante — coluna participant_id e nome dos ficheiros;
  (2) experimento — 1 ou 2 (coluna experiment; condições em conditions/conditions_exp1.csv vs conditions_exp2.csv);
  (3) sessão — Exp1: 1–5; Exp2: 1–10 (ficheiro conditions_exp2_sessions.csv);
  (4) repetição — 1=1ª vez, 2=2ª … nessa condição (coluna session_run);
  (5) Exp1 — tempo máx. de procura do T (s), por trial: diálogo, --search-max-s, FORR_SEARCH_MAX_S
      ou forrageamento_prefs.json; checkbox “Salvar” grava novo padrão (predef. 15);
  (6) Change over delay / pausa COD (ms): diálogo, --cod-grey-ms ou FORR_COD_GREY_MS (predef. 400);
  (7) ponto do olhar: checkbox no diálogo (padrão ligado), --no-gaze-dot ou FORR_SHOW_GAZE_DOT=0;
  (7b) Exp1 — tempo total de procura na ecrã ←/→: checkbox (padrão desligado),
      --show-search-time-hud ou FORR_SHOW_SEARCH_TIME_HUD=1;
  (8) Exp2 — tempo total (s) após instruções: diálogo, --exp2-duration-s, FORR_EXP2_DURATION_S
      ou forrageamento_prefs.json; checkbox “Salvar” grava novo padrão;
  (9) Exp2 — distratores L esquerda/direita: diálogo (pré-preenchidos da condição) ou --n-l-left / --n-l-right.

Exp1 não tem tempo total de sessão: a sessão termina quando todos os trials da lista estiverem
concluídos (ou com T+Esc). search_max_s é um limite **por trial**, só para a procura do T: arranca
quando o olhar fica centrado (drift_ok) e para ao passar à pergunta lateral — inclui a escolha de
metade, o forrageamento e as pausas COD; exclui instruções, fixação central e a resposta ←/→.
Exp2 mantém a duração total configurável (após instruções).

Colunas fixas no CSV (ordem): participant_id, experiment … left_panel, right_panel, target_side, trial, trial_id,
target_t_x, target_t_y, stim_layout_seed, condition_trial, … forage_time_*, cod_switch_count,
cod_grey_ms, search_max_s, search_time_s, target_found, search_timed_out, correctkey, discrete_key, accuracy, … (sem coluna alvos). A ordem dos trials é permutada por sessão
(reproduzível); o layout do S+ depende do participante, sessão, repetição e linha de condição.

Ficheiros em data/: forrageamento_exp<1|2>_<id>_s<sess>_r<rep>_<timestamp>.*

Saídas de sessão (pasta data/): três níveis — (1) trials em .csv; (2) segmentos de forrageamento
em _segments.csv; (3) linha de tempo de eventos em _events.csv; (4) camada _analysis_* (proporções,
matching law, decisão) sem alterar os CSV brutos. TXT junta os três blocos TSV;
XLSX (openpyxl) com folhas trials, segments, events. Modo EyeLink: .EDF após receiveDataFile;
  mensagens !V IAREA RECTANGLE (+ IA_META / IA_BLOCK) alinham IAs do Data Viewer às mesmas
  regiões que o código (coordenadas de ecrã = DISPLAY_COORDS / screen_pixel_coords); IA 110 = rect
  em torno do T (confirmação: olhar dentro + Espaço).

CSV inclui first_side_choice, forage_time_left_s / forage_time_right_s
(tempo com array visível por lado, sem o intervalo COD sem estímulos), cod_switch_count, cod_grey_ms, search_max_s,
search_time_s (desde olhar centrado até achar o T ou timeout da procura), target_found, search_timed_out, accuracy, RT, etc.
Exp1: após pergunta lateral, ecrã cheio 1 s (verde se correto, cinza se errado).
A fixação central (antes do drift_ok) não entra em search_time_s; a resposta lateral usa rt_discrete_s.

Durante a tarefa (instruções e trials): **Esc** sozinho não encerra. **T** e, em seguida (até ~5 s),
**Esc** pede fim da sessão; no ecrã final «Experimento encerrado…» deve premir **Espaço** e
só então os dados são gravados e o programa fecha. Com todos os trials concluídos, o mesmo ecrã final.
Antes do primeiro trial, o menu E/M ainda aceita **Esc** para cancelar sem gravar.

Com EyeLink activo: **X** ou **C** pausam e, após **Espaço**, abrem a recalibração / validação no host
(fluxo semelhante a outros protocolos PsychoPy no laboratório). Em modo dummy (sem tracker), estas teclas não têm efeito.
"""
import argparse
import csv
import hashlib
import json
import math
import os
import random
import sys
import time
from typing import Any, Dict, List, Optional, Tuple

import pylink

try:
    from psychopy import core, event, visual  # noqa: E402
except SyntaxError as _syn_err:
    if sys.version_info < (3, 7):
        sys.stderr.write(
            "\n*** Ambiente incompatível com Python {0}.{1}\n"
            "    Ao importar PsychoPy/visual, uma dependência (pandas, pytest,\n"
            "    cryptography, etc.) usa sintaxe que só existe no Python 3.7+.\n"
            "    Isto acontece quando o PsychoPy 2021 (Py 3.6) recebeu\n"
            "    'pip install' de pacotes demasiado recentes.\n\n"
            "    Solução recomendada: correr este script com o **python.exe**\n"
            "    do **PsychoPy 2023.2.1** (Python 3.8+), em Coder ou terminal.\n\n"
            "    Alternativa: reinstalar o PsychoPy 2021.2.3 standalone oficial\n"
            "    sem atualizar o site-packages, ou usar um venv limpo só para o laboratório.\n\n"
            "    Erro: {2}\n".format(
                sys.version_info[0],
                sys.version_info[1],
                _syn_err,
            )
        )
        raise SystemExit(1)
    raise

try:
    from psychopy.hardware.keyboard import Keyboard as _PsychoKeyboard
except ImportError:
    _PsychoKeyboard = None

from EyeLinkCoreGraphicsPsychoPy import EyeLinkCoreGraphicsPsychoPy

if sys.version_info < (3, 7):
    sys.stderr.write(
        "Aviso: Python {0}.{1}; use PsychoPy 2023 se houver erro ao importar "
        "(p.ex. pandas). PsychoPy 2021.2.3 com Python 3.6 costuma funcionar.\n".format(
            sys.version_info[0], sys.version_info[1]
        )
    )


class _LegacyKeyEvent(object):
    __slots__ = ("name",)

    def __init__(self, name):
        self.name = name


class LegacyKeyboardAdapter(object):
    """Teclas via event.getKeys (PsychoPy antigo sem psychopy.hardware.keyboard)."""

    def clearEvents(self):
        try:
            event.clearEvents(eventType="keyboard")
        except TypeError:
            event.clearEvents()

    def getKeys(self, keyList=None, waitRelease=False):
        try:
            keys = event.getKeys(keyList=keyList, timeStamped=False)
        except TypeError:
            keys = event.getKeys(keyList=keyList)
        return [_LegacyKeyEvent(k) for k in keys]


def make_keyboard():
    if _PsychoKeyboard is not None:
        return _PsychoKeyboard()
    return LegacyKeyboardAdapter()


def flush_keyboard(kb=None):
    """Evita que teclas pendentes (ex.: ao carregar no Coder) saltem instruções."""
    try:
        if kb is not None:
            kb.clearEvents()
    except Exception:
        pass
    try:
        event.clearEvents(eventType="keyboard")
    except TypeError:
        event.clearEvents()
    for _ in range(5):
        event.getKeys()
        core.wait(0.02)


def run_mode_selection(win: visual.Window, kb, eyelink_ip: str):
    """
    Menu inicial: E = EyeLink, M/D = rato (dummy). ESC = abortar sessão.
    Retorna True se dummy, False se EyeLink, None se saiu (ESC),
    ou a string "quit_save" se T+Esc (gravar e sair).
    """
    flush_keyboard(kb)
    core.wait(0.15)
    stim = visual.TextStim(
        win,
        text=(
            "Escolha o modo antes de iniciar:\n\n"
            "   E   —   EyeLink  (host: {0})\n"
            "   M   —   Rato como olhar  (dummy, sem tracker)\n"
            "   D   —   Igual a M  (dummy)\n\n"
            "   ESC —   Sair\n"
            "   T depois Esc —   Gravar dados e sair"
        ).format(eyelink_ip),
        wrapWidth=win.size[0] * 0.85,
        height=max(22.0, min(30.0, win.size[1] * 0.024)),
        color="white",
        units="pix",
        bold=True,
        font=INSTRUCTIONS_UI_FONT,
    )
    while True:
        poll_global_quit_t_esc(kb, None, win)
        if QUIT_SAVE_REQUESTED:
            return "quit_save"
        if "escape" in event.getKeys():
            return None
        keys = kb.getKeys(
            keyList=["e", "m", "d", "escape"],
            waitRelease=False,
        )
        if keys:
            name = keys[0].name
            if name == "escape":
                return None
            if name in ("m", "d"):
                return True
            if name == "e":
                return False
        stim.draw()
        win.flip()


def show_error_screen(win: visual.Window, kb, title: str, detail: str):
    """Mostra mensagem em fullscreen antes de fechar (evita 'abre e fecha' sem perceber o motivo)."""
    flush_keyboard(kb)
    msg = visual.TextStim(
        win,
        text=title + "\n\n" + detail + "\n\nPressione ESPAÇO para fechar.",
        wrapWidth=win.size[0] * 0.85,
        height=max(20.0, min(28.0, win.size[1] * 0.022)),
        color="white",
        units="pix",
        font=INSTRUCTIONS_UI_FONT,
    )
    while True:
        keys = kb.getKeys(keyList=["space", "escape"], waitRelease=False)
        if keys:
            break
        msg.draw()
        win.flip()


def run_experiment_end_screen(win: visual.Window, kb, tracker: Any) -> None:
    """
    Ecrã final: após Espaço o programa grava dados e fecha (ver main).
    Durante a tarefa só T+Esc ou concluir todos os trials leva aqui.
    """
    send_msg(tracker, "PHASE experiment_end_screen")
    flush_keyboard(kb)
    sw, sh = float(win.size[0]), float(win.size[1])
    sc = layout_scale(sw, sh)
    msg = visual.TextStim(
        win,
        text=(
            "Experimento encerrado.\n\n"
            "Chame o experimentador.\n\n"
            "Aperte Espaço para sair."
        ),
        pos=(0.0, 0.0),
        height=max(18.0, 24.0 * sc),
        color=(0.9, 0.9, 0.9),
        units="pix",
        wrapWidth=sw * 0.82,
        bold=True,
        font=INSTRUCTIONS_UI_FONT,
    )
    while True:
        keys = kb.getKeys(keyList=["space"], waitRelease=False)
        if keys:
            break
        win.color = BACKGROUND_COLOR
        msg.draw()
        win.flip()
    flush_keyboard(kb)


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
_THIS_DIR = os.path.abspath(os.path.dirname(__file__))
_CONDITIONS_FILE_EXP1 = os.path.join(_THIS_DIR, "conditions", "conditions_exp1.csv")
_CONDITIONS_FILE_EXP2 = os.path.join(_THIS_DIR, "conditions", "conditions_exp2.csv")
_CONDITIONS_FILE_EXP2_SESSIONS = os.path.join(
    _THIS_DIR, "conditions", "conditions_exp2_sessions.csv"
)
_CONDITIONS_FILE = _CONDITIONS_FILE_EXP1  # compatibilidade com código antigo
_PREFS_FILE = os.path.join(_THIS_DIR, "forrageamento_prefs.json")
_DATA_FOLDER = os.path.join(_THIS_DIR, "data")

EXPERIMENT_MIN = 1
EXPERIMENT_MAX = 2

# ---------------------------------------------------------------------------
# Settings (protocolo / EB)
# ---------------------------------------------------------------------------
BACKGROUND_COLOR = (-0.4, -0.4, -0.4)
# Quadrado no lado “inativo”: mais escuro que o fundo (visível sem PNG).
SIDE_MARKER_FILL = (-0.72, -0.72, -0.72)
SIDE_MARKER_LINE = (-0.88, -0.88, -0.88)
# Exp2: mesmo marcador, tom mais claro que SIDE_MARKER_* mas ainda ≠ BACKGROUND_COLOR.
EXP2_INACTIVE_SIDE_FILL = (-0.52, -0.52, -0.52)
EXP2_INACTIVE_SIDE_LINE = (-0.64, -0.64, -0.64)
DRIFT_FIX_MS = 300
REGION_CHOICE_MIN_DWELL_MS = 200
PRE_CHOICE_HOLD_MS = 800
# Faixa vertical central na escolha de metade: olhar aqui não inicia escolha; contorno branco no ecrã.
REGION_CHOICE_NEUTRAL_HALF_W_FRAC = 0.09
REGION_CHOICE_NEUTRAL_HALF_W_MIN_PX = 72.0
COD_FIX_MS = 600
# Pausa COD ao trocar de alternativa (fundo = BACKGROUND_COLOR, sem painéis); configurável.
DEFAULT_COD_GREY_MS = 400
COD_GREY_MS = DEFAULT_COD_GREY_MS
TARGET_FIX_MS = 500
# Exp1: limite de procura do T (s) **por trial**, contado só na fase de procura (desde
# foraging_start, depois da fixação central e da escolha de metade); diálogo / --search-max-s.
DEFAULT_SEARCH_MAX_S = 15.0
CORRECT_FEEDBACK_S = 1.0
CORRECT_FEEDBACK_COLOR = (-0.35, 0.78, -0.35)
SEARCH_MAX_SAVE_DLG_KEY = "Salvar"
COD_GREY_MS_DLG_KEY = 'Change over delay (ms) — 400 ("Padrão")'
SHOW_GAZE_DOT_DLG_KEY = "Mostrar ponto do olhar"
DEFAULT_SHOW_GAZE_DOT = True
# Exp1: HUD com soma do tempo de procura (search_time_s) na ecrã ←/→; padrão desligado.
SHOW_SEARCH_TIME_HUD_DLG_KEY = "Mostrar tempo total de procura (Exp. 1)"
DEFAULT_SHOW_SEARCH_TIME_HUD = False
# Exp2: duração da sessão de forrageamento (s), após instruções; prefs / diálogo / CLI.
DEFAULT_EXP2_DURATION_S = 420.0
EXP2_DURATION_SAVE_DLG_KEY = "Salvar"
N_L_LEFT_DLG_KEY = "Distratores L — esquerda (Exp. 2)"
N_L_RIGHT_DLG_KEY = "Distratores L — direita (Exp. 2)"

# Referência de layout (EB antigo); escala com layout_scale(sw, sh)
REF_LAYOUT_W = 1280.0
REF_LAYOUT_H = 768.0
# Tamanho lógico da janela em modo janela cheia (o monitor pode ter outra resolução nativa)
DEFAULT_WINDOW_W = 1680
DEFAULT_WINDOW_H = 1050
# Distância olhos–ecrã (protocolo EB / método). Registo em mensagens EyeLink e cabeçalhos de sessão.
VIEWING_DISTANCE_CM = 70
# Presets para --resolution / FORR_RESOLUTION / diálogo (apenas strings "WxH")
RESOLUTION_PRESET_CHOICES = (
    "1680x1050",
    "1920x1080",
    "1366x768",
    "1280x1024",
)
DRIFT_AOI_RADIUS = 60
# Confirmação do alvo: AOI em torno do glifo T (centro = pos do TextStim). Espaço só avança com olhar
# dentro desta caixa + TARGET_FIX_MS na metade ativa onde está o T (alinhado à IA 110 no EDF).
# Meias-dimensões face ao original (0.78 / 0.88 / 22): ×1.5 ×1.2 ×1.2 ×1.07 ×1.25 ×0.9 ×1.05 — Exp1 e Exp2.
TARGET_T_AOI_HALF_W_FRAC = 2.1294819
TARGET_T_AOI_HALF_H_FRAC = 2.4024924
TARGET_T_AOI_MIN_HALF_PX = 60.06231
# Margem dentro de cada metade do ecrã (px na resolução de referência ~ proporcional)
LAYOUT_MARGIN_X = 16
LAYOUT_MARGIN_Y = 24
# Proporção do painel de forrageamento (largura × altura em unidades relativas)
PANEL_NAT_FALLBACK = (630, 800)
# Densidade de letras no painel de busca: 0.25 = 1/3 da densidade anterior (0.75);
# ~1/3 dos itens por alternativa; 1 T mantém-se no lado alvo.
SEARCH_PANEL_STIM_COUNT_FRAC = 0.25
SEARCH_PANEL_STIM_COUNT_MIN = 20

# Tons de cinza (RGB -1…1) para L’s e T (sorteio independente por letra)
GRAYS_FOR_L = (
    (-0.22, -0.22, -0.22),
    (-0.12, -0.12, -0.12),
    (-0.02, -0.02, -0.02),
    (0.08, 0.08, 0.08),
    (0.18, 0.18, 0.18),
    (0.28, 0.28, 0.28),
    (0.38, 0.38, 0.38),
)
FIXATION_GRAY = (0.45, 0.45, 0.45)
COD_INDICATOR_GRAY = (0.35, 0.35, 0.35)
# Marcador de olhar: um único círculo, com a mesma área visual nas duas metades.
# A posição mostrada é exatamente a amostra usada pelas regras de dwell/AOI.
GAZE_DOT_RADIUS_PX = 4.0
GAZE_DOT_FILL = (0.92, 0.92, 0.92)
GAZE_DOT_LINE = (-0.72, -0.72, -0.72)
# T alvo só no painel de exemplo das instruções (verde; na tarefa usa GRAYS_FOR_L)
INSTRUCTIONS_PREVIEW_TARGET_GREEN = (-0.92, 0.62, -0.55)
INSTRUCTIONS_PREVIEW_BOX_FILL = (-0.28, -0.28, -0.28)
INSTRUCTIONS_PREVIEW_BOX_LINE = (-0.75, -0.75, -0.75)
# Mesma família tipográfica no título e no corpo das instruções (Windows: Arial).
INSTRUCTIONS_UI_FONT = "Arial"


def instruction_font_heights(sh: float) -> Tuple[float, float, float]:
    """Alturas do glifo (pix) para título, corpo e rodapé das instruções."""
    return (
        max(20.0, min(34.0, sh * 0.034)),
        max(14.0, min(21.0, sh * 0.021)),
        max(15.0, min(24.0, sh * 0.024)),
    )


def instruction_example_panel_size(
    sw: float, sh: float, frac: float = 0.30
) -> Tuple[float, float]:
    """Largura e altura (pix) da caixa L/T nas instruções (proporção do lado menor do ecrã)."""
    w = min(float(sw), float(sh)) * float(frac)
    return w, w * 0.92


def make_instruction_text(
    win: visual.Window,
    text: str,
    y_pos: float,
    height: float,
    wrap_width: float,
    color: Tuple[float, float, float] = (1.0, 1.0, 1.0),
) -> visual.TextStim:
    return visual.TextStim(
        win,
        text=text,
        pos=(0.0, y_pos),
        height=height,
        color=color,
        units="pix",
        wrapWidth=wrap_width,
        bold=True,
        font=INSTRUCTIONS_UI_FONT,
    )

EDF_FILENAME = "FVEXP1.EDF"

# Só usado com --skip-mode-menu: ignora EyeLink sem --dummy (legado).
RUN_DUMMY_WHEN_NO_ARGS = False

# T → Esc: encerrar sessão e gravar CSV (ver poll_global_quit_t_esc)
T_ESC_ARM_TIMEOUT_S = 5.0
_t_esc_arm_until = 0.0
QUIT_SAVE_REQUESTED = False

# IP do PC EyeLink (rede local). Ex.: "192.168.1.100". Igual ao "Tracker Address" no Experiment Builder.
DEFAULT_EYELINK_HOST = os.environ.get("FORR_EYELINK_HOST", "100.1.1.1")

# Cinco níveis de condição experimental (1–5) no Exp1; Exp2 usa sessões 1–10 no CSV dedicado.
SESSION_CONDITION_MIN = 1
SESSION_CONDITION_MAX = 5
SESSION_CONDITION_MAX_EXP2 = 10
# Exp2: total de L distratores (ambos painéis); decaimento simétrico k L/s/painel (5×75/60).
EXP2_N_DISTRACTOR_TOTAL = 150
EXP2_DECAY_REF_PER_S = 5
EXP2_DECAY_REF_MAX_PER_SIDE = 60


def session_condition_max_for_experiment(experiment_number: int) -> int:
    return (
        SESSION_CONDITION_MAX_EXP2
        if int(experiment_number) == 2
        else SESSION_CONDITION_MAX
    )


def exp2_decay_k_per_panel() -> int:
    """Remoções de L por painel e por segundo (simétrico)."""
    return max(
        1,
        int(
            round(
                float(EXP2_DECAY_REF_PER_S)
                * (float(EXP2_N_DISTRACTOR_TOTAL) / 2.0)
                / float(EXP2_DECAY_REF_MAX_PER_SIDE)
            )
        ),
    )


def trial_order_seed(
    participant_id: str,
    experiment: int,
    session_condition: int,
    session_run: int,
) -> int:
    """Semente determinística para permutar a ordem dos trials (reproduzível por sessão)."""
    h = hashlib.sha256(
        "trial_order|{0}|{1}|{2}|{3}".format(
            participant_id,
            int(experiment),
            int(session_condition),
            int(session_run),
        ).encode("utf-8")
    ).digest()
    return int.from_bytes(h[:8], "big")


def stimulus_layout_seed(
    participant_id: str,
    experiment: int,
    session_condition: int,
    session_run: int,
    presentation_index: int,
    condition_trial_key: str,
) -> int:
    """
    Semente para RNG dos painéis L/T: a posição do S+ varia com participante, experiência,
    sessão, repetição, ordem de apresentação e linha de condição.
    """
    h = hashlib.sha256(
        "stim_layout|{0}|{1}|{2}|{3}|{4}|{5}".format(
            participant_id,
            int(experiment),
            int(session_condition),
            int(session_run),
            int(presentation_index),
            (condition_trial_key or "").strip(),
        ).encode("utf-8")
    ).digest()
    return int.from_bytes(h[:8], "big")


def panels_lt_from_target_side(target_side: str) -> Tuple[str, str]:
    """Painel esquerdo/direito: T onde está o S+; L no outro."""
    ts = (target_side or "").strip().lower()
    if ts == "left":
        return ("T", "L")
    return ("L", "T")


def make_valid_edf_name(participant_id: str) -> str:
    """Nome EDF no host EyeLink: máx. 8 caracteres alfanuméricos + .EDF"""
    s = "".join(c for c in str(participant_id).upper() if c.isalnum())
    if not s:
        s = "FVEXP1"
    return s[:8] + ".EDF"


def sanitize_participant_id(raw: str) -> str:
    """Identificador seguro para nomes de ficheiro (Windows)."""
    s = str(raw).strip()
    bad = '<>:"/\\|?*'
    s = "".join((c if c not in bad else "_") for c in s)
    s = "".join(c if (c.isalnum() or c in "_- ") else "_" for c in s)
    s = "_".join(s.split())
    s = s.strip("_") or "SEMID"
    return s[:64]


def parse_wxh_string(raw: Any) -> Optional[Tuple[int, int]]:
    """Interpreta '1680x1050', '1680 x 1050', '1680×1050' → (largura, altura)."""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if isinstance(raw, list) and raw:
        s = str(raw[0]).strip().lower()
    if not s:
        return None
    for ch in ("\u00d7", "\u2715"):
        s = s.replace(ch, "x")
    s = "".join(c for c in s if not c.isspace())
    i = s.find("x")
    if i < 1:
        return None
    try:
        w = int(s[:i])
        h = int(s[i + 1 :])
    except ValueError:
        return None
    if w < 160 or h < 120:
        return None
    return (w, h)


def resolve_window_size(
    args: Any,
    gui_w: Optional[int],
    gui_h: Optional[int],
) -> Tuple[int, int]:
    """
    Combina CLI, ambiente e escolha do diálogo. Precedência quando não há ambos --window-width/--height:
    predefinição ou diálogo; depois `--resolution` (sempre); `FORR_RESOLUTION` só se não houver
    valores do diálogo; por fim FORR_WIN_* e --window-* individuais.
    """
    miw, mih = 320, 240
    w_cli = getattr(args, "window_width", None)
    h_cli = getattr(args, "window_height", None)
    if w_cli is not None and h_cli is not None:
        return max(miw, int(w_cli)), max(mih, int(h_cli))

    w, h = DEFAULT_WINDOW_W, DEFAULT_WINDOW_H
    if gui_w is not None and gui_h is not None:
        w, h = int(gui_w), int(gui_h)

    cli_res = (getattr(args, "resolution", None) or "").strip()
    if cli_res:
        pr_cli = parse_wxh_string(cli_res.replace(" ", ""))
        if pr_cli:
            w, h = pr_cli
    else:
        gui_missing = gui_w is None or gui_h is None
        env_res = (os.environ.get("FORR_RESOLUTION") or "").strip()
        if gui_missing and env_res:
            pr_env = parse_wxh_string(env_res.replace(" ", ""))
            if pr_env:
                w, h = pr_env

    ews = os.environ.get("FORR_WIN_W", "").strip()
    ehs = os.environ.get("FORR_WIN_H", "").strip()
    if ews.isdigit():
        w = int(ews)
    if ehs.isdigit():
        h = int(ehs)

    if w_cli is not None:
        w = max(miw, int(w_cli))
    if h_cli is not None:
        h = max(mih, int(h_cli))

    return max(miw, int(w)), max(mih, int(h))


def _resolution_dropdown_choices(default_w: int, default_h: int) -> List[str]:
    key = "{0}x{1}".format(int(default_w), int(default_h))
    presets = [p for p in RESOLUTION_PRESET_CHOICES if p != key]
    return [key] + presets


def _session_choice_list(default_session: int, experiment_number: int = 1) -> List[str]:
    """Lista de sessões com o valor por defeito em primeiro (DlgFromDict usa o 1.º)."""
    smax = session_condition_max_for_experiment(experiment_number)
    opts = [str(i) for i in range(SESSION_CONDITION_MIN, smax + 1)]
    d = max(SESSION_CONDITION_MIN, min(smax, int(default_session)))
    i0 = d - 1
    return opts[i0:] + opts[:i0]


def _experiment_choice_list(default_experiment: int) -> List[str]:
    opts = ["Experimento 1", "Experimento 2"]
    if default_experiment == 2:
        return ["Experimento 2", "Experimento 1"]
    return opts


def _validate_search_max_s(seconds: float) -> float:
    """Segundos > 0 para o limite de procura do T (por trial) no Experimento 1."""
    s = float(seconds)
    if s <= 0.0:
        raise ValueError("tempo de procura deve ser > 0")
    return s


def get_search_max_s_padrao() -> float:
    """Padrão gravado em prefs, senão DEFAULT_SEARCH_MAX_S (15)."""
    raw = _load_prefs().get("search_max_s", None)
    if raw is None:
        return float(DEFAULT_SEARCH_MAX_S)
    try:
        return _validate_search_max_s(float(raw))
    except (ValueError, TypeError):
        return float(DEFAULT_SEARCH_MAX_S)


def save_search_max_s_padrao(seconds: float) -> float:
    """Grava novo padrão de tempo máx. de procura por trial (Exp1)."""
    v = _validate_search_max_s(seconds)
    prefs = _load_prefs()
    prefs["search_max_s"] = v
    _save_prefs(prefs)
    return v


def search_max_s_dlg_key(padrao_s: float) -> str:
    n = int(padrao_s) if float(padrao_s) == int(padrao_s) else padrao_s
    return 'Tempo máx. procura do T por trial (s) — {0} ("Padrão")'.format(n)


def _default_search_max_s_from_env() -> float:
    raw = os.environ.get("FORR_SEARCH_MAX_S", "").strip()
    if raw:
        try:
            return _validate_search_max_s(float(raw.replace(",", ".")))
        except ValueError:
            pass
    return get_search_max_s_padrao()


def resolve_search_max_seconds(args: Any, experiment_number: int) -> float:
    """Limite de procura do T por trial (s); só aplicado no Exp1 em run_trial."""
    if experiment_number != 1:
        return float(DEFAULT_SEARCH_MAX_S)
    cli = getattr(args, "search_max_s", None)
    if cli is not None:
        try:
            return _validate_search_max_s(float(cli))
        except ValueError:
            print("--search-max-s deve ser um número > 0.", file=sys.stderr)
            raise
    return _default_search_max_s_from_env()


def _validate_cod_grey_ms(ms: float) -> float:
    """Milissegundos >= 0 para a pausa COD (change over delay)."""
    v = float(ms)
    if v < 0.0:
        raise ValueError("change over delay deve ser >= 0")
    return v


def _default_cod_grey_ms_from_env() -> float:
    raw = os.environ.get("FORR_COD_GREY_MS", "").strip()
    if not raw:
        return float(DEFAULT_COD_GREY_MS)
    try:
        return _validate_cod_grey_ms(float(raw.replace(",", ".")))
    except ValueError:
        return float(DEFAULT_COD_GREY_MS)


def resolve_cod_grey_ms(args: Any) -> float:
    """Pausa COD (ms) ao trocar de alternativa; Exp1 e Exp2."""
    cli = getattr(args, "cod_grey_ms", None)
    if cli is not None:
        try:
            return _validate_cod_grey_ms(float(cli))
        except ValueError:
            print("--cod-grey-ms deve ser um número >= 0.", file=sys.stderr)
            raise
    return _default_cod_grey_ms_from_env()


def _parse_dlg_bool(raw: Any, default: bool = True) -> bool:
    if isinstance(raw, bool):
        return raw
    if raw is None:
        return bool(default)
    s = str(raw).strip().lower()
    if s in ("0", "false", "no", "off", "não", "nao"):
        return False
    if s in ("1", "true", "yes", "on", "sim"):
        return True
    return bool(default)


def _default_show_gaze_dot_from_env() -> bool:
    raw = os.environ.get("FORR_SHOW_GAZE_DOT", "").strip().lower()
    if not raw:
        return bool(DEFAULT_SHOW_GAZE_DOT)
    if raw in ("0", "false", "no", "off"):
        return False
    if raw in ("1", "true", "yes", "on"):
        return True
    return bool(DEFAULT_SHOW_GAZE_DOT)


def resolve_show_gaze_dot(args: Any) -> bool:
    """Se True, desenha o ponto que segue o olhar/rato (padrão: ligado)."""
    if getattr(args, "no_gaze_dot", False):
        return False
    cli = getattr(args, "show_gaze_dot", None)
    if cli is not None:
        return bool(cli)
    return _default_show_gaze_dot_from_env()


def _default_show_search_time_hud_from_env() -> bool:
    raw = os.environ.get("FORR_SHOW_SEARCH_TIME_HUD", "").strip().lower()
    if not raw:
        return bool(DEFAULT_SHOW_SEARCH_TIME_HUD)
    if raw in ("0", "false", "no", "off"):
        return False
    if raw in ("1", "true", "yes", "on"):
        return True
    return bool(DEFAULT_SHOW_SEARCH_TIME_HUD)


def resolve_show_search_time_hud(args: Any) -> bool:
    """Exp1: se True, mostra na ecrã ←/→ a soma dos search_time_s da sessão (padrão: off)."""
    if getattr(args, "show_search_time_hud", False):
        return True
    return _default_show_search_time_hud_from_env()


def format_search_time_hud_text(total_s: float) -> str:
    """Texto do HUD de tempo total de procura (soma dos trials)."""
    t = max(0.0, float(total_s))
    return "Tempo total de procura: {0:.1f} s".format(t)


def _load_prefs() -> Dict[str, Any]:
    try:
        with open(_PREFS_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            return data
    except (OSError, ValueError, TypeError):
        pass
    return {}


def _save_prefs(prefs: Dict[str, Any]) -> None:
    with open(_PREFS_FILE, "w", encoding="utf-8") as f:
        json.dump(prefs, f, indent=2, ensure_ascii=False)
        f.write("\n")


def _validate_exp2_duration_s(seconds: float) -> float:
    s = float(seconds)
    if s <= 0.0:
        raise ValueError("tempo total Exp2 deve ser > 0")
    return s


def get_exp2_duration_padrao() -> float:
    """Padrão gravado em prefs, senão DEFAULT_EXP2_DURATION_S (420)."""
    raw = _load_prefs().get("exp2_duration_s", None)
    if raw is None:
        return float(DEFAULT_EXP2_DURATION_S)
    try:
        return _validate_exp2_duration_s(float(raw))
    except (ValueError, TypeError):
        return float(DEFAULT_EXP2_DURATION_S)


def save_exp2_duration_padrao(seconds: float) -> float:
    """Grava novo padrão de duração Exp2 e devolve o valor validado."""
    v = _validate_exp2_duration_s(seconds)
    prefs = _load_prefs()
    prefs["exp2_duration_s"] = v
    _save_prefs(prefs)
    return v


def exp2_duration_dlg_key(padrao_s: float) -> str:
    n = int(padrao_s) if float(padrao_s) == int(padrao_s) else padrao_s
    return 'Tempo total Exp. 2 (s) — {0} ("Padrão")'.format(n)


def _default_exp2_duration_s_from_env_or_prefs() -> float:
    raw = os.environ.get("FORR_EXP2_DURATION_S", "").strip()
    if raw:
        try:
            return _validate_exp2_duration_s(float(raw.replace(",", ".")))
        except ValueError:
            pass
    return get_exp2_duration_padrao()


def resolve_exp2_duration_s(args: Any, experiment_number: int = 2) -> float:
    """Duração (s) da sessão Exp2 após instruções. Irrelevante no Exp1 (devolve o padrão)."""
    if int(experiment_number) != 2:
        return float(DEFAULT_EXP2_DURATION_S)
    cli = getattr(args, "exp2_duration_s", None)
    if cli is not None:
        try:
            return _validate_exp2_duration_s(float(cli))
        except ValueError:
            print("--exp2-duration-s deve ser um número > 0.", file=sys.stderr)
            raise
    return _default_exp2_duration_s_from_env_or_prefs()


def _validate_n_l_count(n: float) -> int:
    v = int(round(float(n)))
    if v < 0:
        raise ValueError("número de L deve ser >= 0")
    return v


def load_exp2_condition_n_l(session: int) -> Tuple[int, int, str]:
    """
    Lê n_L_left, n_L_right e ratio_label da condição (sessão) no CSV Exp2.
    Fallback: 75/75 e rótulo vazio.
    """
    path = _CONDITIONS_FILE_EXP2_SESSIONS
    try:
        with open(path, newline="", encoding="utf-8") as f:
            for row in csv.DictReader(f):
                try:
                    if int((row.get("session") or "0").strip()) == int(session):
                        nl = _validate_n_l_count(float(row.get("n_L_left") or 75))
                        nr = _validate_n_l_count(float(row.get("n_L_right") or 75))
                        ratio = (row.get("ratio_label") or "").strip()
                        return nl, nr, ratio
                except (ValueError, TypeError):
                    continue
    except OSError:
        pass
    return 75, 75, ""


def resolve_exp2_n_l(
    args: Any, session: int, experiment_number: int
) -> Tuple[Optional[int], Optional[int]]:
    """Contagens iniciais de L por lado no Exp2 (CSV da condição, ou CLI)."""
    if int(experiment_number) != 2:
        return None, None
    csv_l, csv_r, _ = load_exp2_condition_n_l(session)
    cli_l = getattr(args, "n_l_left", None)
    cli_r = getattr(args, "n_l_right", None)
    try:
        left = _validate_n_l_count(cli_l) if cli_l is not None else csv_l
        right = _validate_n_l_count(cli_r) if cli_r is not None else csv_r
    except ValueError:
        print("--n-l-left / --n-l-right devem ser inteiros >= 0.", file=sys.stderr)
        raise
    return left, right


def _parse_experiment_field(raw: Any) -> int:
    if isinstance(raw, list) and raw:
        raw = raw[0]
    s = str(raw).strip()
    if s == "Experimento 2":
        return 2
    return 1


def _ask_exp2_only_params_interactive(
    session: int,
    default_duration_s: float,
) -> Optional[Tuple[float, Optional[int], Optional[int]]]:
    """
    Diálogo só com parâmetros do Exp2 (duração + Salvar + L), usado quando o
    formulário principal abriu como Exp1 e o utilizador escolheu Exp2.
    Devolve (duration_s, n_L_left, n_L_right) ou None se cancelar.
    """
    try:
        from psychopy import gui
    except Exception:
        return (
            float(default_duration_s),
            load_exp2_condition_n_l(session)[0],
            load_exp2_condition_n_l(session)[1],
        )

    padrao = get_exp2_duration_padrao()
    dur_key = exp2_duration_dlg_key(padrao)
    csv_l, csv_r, ratio = load_exp2_condition_n_l(session)
    left_key = N_L_LEFT_DLG_KEY
    right_key = N_L_RIGHT_DLG_KEY
    if ratio:
        left_key = "Distratores L — esquerda ({0})".format(ratio)
        right_key = "Distratores L — direita ({0})".format(ratio)
    def_dur = float(default_duration_s)
    info = {
        dur_key: str(int(def_dur) if def_dur == int(def_dur) else def_dur),
        EXP2_DURATION_SAVE_DLG_KEY: False,
        left_key: str(csv_l),
        right_key: str(csv_r),
    }
    dlg = gui.DlgFromDict(
        dictionary=info,
        title="Experimento 2 — tempo e distratores",
        order=[dur_key, EXP2_DURATION_SAVE_DLG_KEY, left_key, right_key],
    )
    if not dlg.OK:
        return None
    try:
        duration_s = _validate_exp2_duration_s(
            float(str(info.get(dur_key) or def_dur).strip().replace(",", "."))
        )
    except ValueError:
        print(
            "Tempo total Exp. 2 inválido (número > 0, em segundos; conta após instruções).",
            file=sys.stderr,
        )
        return None
    if _parse_dlg_bool(info.get(EXP2_DURATION_SAVE_DLG_KEY), default=False):
        try:
            saved = save_exp2_duration_padrao(duration_s)
            print(
                'Novo padrão Exp. 2 gravado: {0:g} s (após instruções) → {1}'.format(
                    saved, _PREFS_FILE
                )
            )
        except (OSError, ValueError) as exc:
            print(
                "Aviso: não foi possível gravar o padrão Exp. 2 ({0}).".format(exc),
                file=sys.stderr,
            )
    try:
        n_L_left = _validate_n_l_count(
            float(str(info.get(left_key) or csv_l).strip().replace(",", "."))
        )
        n_L_right = _validate_n_l_count(
            float(str(info.get(right_key) or csv_r).strip().replace(",", "."))
        )
    except ValueError:
        print(
            "Distratores L inválidos (inteiros >= 0 para esquerda e direita).",
            file=sys.stderr,
        )
        return None
    return duration_s, n_L_left, n_L_right


def _import_qt_widgets():
    """PyQt via PsychoPy (ou PyQt5 directo). Devolve (QtWidgets, None) ou (None, None)."""
    try:
        from psychopy.gui.qtgui import QtWidgets, ensureQtApp

        ensureQtApp()
        return QtWidgets
    except Exception:
        pass
    try:
        from PyQt5 import QtWidgets

        app = QtWidgets.QApplication.instance()
        if app is None:
            QtWidgets.QApplication([])
        return QtWidgets
    except Exception:
        return None


def _ask_experiment_setup_qt(
    default_participant: str,
    default_session: int,
    default_run: int,
    default_experiment: int,
    default_search_max_s: float,
    default_cod_grey_ms: float,
    default_show_gaze_dot: bool,
    default_exp2_duration_s: float,
    default_show_search_time_hud: bool = False,
) -> Optional[
    Tuple[
        str,
        int,
        int,
        int,
        int,
        int,
        float,
        float,
        bool,
        bool,
        float,
        Optional[int],
        Optional[int],
    ]
]:
    """
    Diálogo Qt com campos condicionais:
      - Sessão + repetição: Exp1 e Exp2
      - Tempo máx. procura + Salvar + HUD tempo total: só Exp1
      - Tempo total + Salvar + L esq/dir: só Exp2 (L atualiza ao mudar sessão)
    Ao mudar Experimento 1↔2, os campos específicos aparecem/desaparecem no mesmo diálogo.
    """
    QtWidgets = _import_qt_widgets()
    if QtWidgets is None:
        return None

    padrao_exp2 = get_exp2_duration_padrao()
    padrao_search = get_search_max_s_padrao()
    de0 = max(EXPERIMENT_MIN, min(EXPERIMENT_MAX, int(default_experiment)))
    ds0 = max(
        SESSION_CONDITION_MIN,
        min(session_condition_max_for_experiment(de0), int(default_session)),
    )
    env_res = parse_wxh_string(os.environ.get("FORR_RESOLUTION", "").strip())
    def_w = env_res[0] if env_res else DEFAULT_WINDOW_W
    def_h = env_res[1] if env_res else DEFAULT_WINDOW_H
    res_choices = _resolution_dropdown_choices(def_w, def_h)

    dlg = QtWidgets.QDialog()
    dlg.setWindowTitle("Forrageamento visual")
    dlg.setModal(True)
    form = QtWidgets.QFormLayout(dlg)

    edit_participant = QtWidgets.QLineEdit((default_participant or "").strip())
    combo_exp = QtWidgets.QComboBox()
    combo_exp.addItems(["Experimento 1", "Experimento 2"])
    combo_exp.setCurrentIndex(0 if de0 == 1 else 1)

    combo_session = QtWidgets.QComboBox()
    label_session = QtWidgets.QLabel()
    edit_run = QtWidgets.QLineEdit(str(max(1, int(default_run))))
    label_run = QtWidgets.QLabel("Repetição desta condição (1=1ª vez, 2=2ª…)")

    # Exp1-only
    label_search = QtWidgets.QLabel()
    sms0 = float(default_search_max_s)
    edit_search = QtWidgets.QLineEdit(
        str(int(sms0) if sms0 == int(sms0) else sms0)
    )
    check_save_search = QtWidgets.QCheckBox("Salvar")
    check_search_hud = QtWidgets.QCheckBox(SHOW_SEARCH_TIME_HUD_DLG_KEY)
    check_search_hud.setChecked(bool(default_show_search_time_hud))

    # Shared
    edit_cod = QtWidgets.QLineEdit(
        str(
            int(default_cod_grey_ms)
            if float(default_cod_grey_ms) == int(default_cod_grey_ms)
            else default_cod_grey_ms
        )
    )
    check_gaze = QtWidgets.QCheckBox(SHOW_GAZE_DOT_DLG_KEY)
    check_gaze.setChecked(bool(default_show_gaze_dot))
    combo_res = QtWidgets.QComboBox()
    combo_res.addItems(res_choices)

    # Exp2-only
    dur_label = QtWidgets.QLabel()
    edit_dur = QtWidgets.QLineEdit()
    check_save_dur = QtWidgets.QCheckBox("Salvar")
    edit_n_l_left = QtWidgets.QLineEdit()
    edit_n_l_right = QtWidgets.QLineEdit()
    label_n_l_left = QtWidgets.QLabel(N_L_LEFT_DLG_KEY)
    label_n_l_right = QtWidgets.QLabel(N_L_RIGHT_DLG_KEY)

    form.addRow("Código do participante", edit_participant)
    form.addRow("Experimento", combo_exp)
    form.addRow(label_session, combo_session)
    form.addRow(label_run, edit_run)
    form.addRow(label_search, edit_search)
    form.addRow(check_save_search)
    form.addRow(check_search_hud)
    form.addRow(dur_label, edit_dur)
    form.addRow(check_save_dur)
    form.addRow(label_n_l_left, edit_n_l_left)
    form.addRow(label_n_l_right, edit_n_l_right)
    form.addRow(COD_GREY_MS_DLG_KEY, edit_cod)
    form.addRow(check_gaze)
    form.addRow("Resolução da janela", combo_res)

    buttons = QtWidgets.QDialogButtonBox(
        QtWidgets.QDialogButtonBox.Ok | QtWidgets.QDialogButtonBox.Cancel
    )
    buttons.accepted.connect(dlg.accept)
    buttons.rejected.connect(dlg.reject)
    form.addRow(buttons)

    def _fill_sessions(exp_n):
        smax = session_condition_max_for_experiment(exp_n)
        label_session.setText("Sessão (condição 1 a {0})".format(smax))
        cur = combo_session.currentText().strip()
        combo_session.blockSignals(True)
        combo_session.clear()
        for i in range(SESSION_CONDITION_MIN, smax + 1):
            combo_session.addItem(str(i))
        idx = combo_session.findText(cur)
        if idx < 0:
            prefer = str(ds0) if exp_n == de0 else "1"
            idx = combo_session.findText(prefer)
        if idx < 0:
            idx = 0
        combo_session.setCurrentIndex(max(0, idx))
        combo_session.blockSignals(False)

    def _apply_session_n_l():
        if combo_exp.currentIndex() != 1:
            return
        try:
            sess = int(combo_session.currentText().strip())
        except ValueError:
            return
        nl, nr, ratio = load_exp2_condition_n_l(sess)
        edit_n_l_left.setText(str(nl))
        edit_n_l_right.setText(str(nr))
        if ratio:
            label_n_l_left.setText(
                "Distratores L — esquerda ({0})".format(ratio)
            )
            label_n_l_right.setText(
                "Distratores L — direita ({0})".format(ratio)
            )
        else:
            label_n_l_left.setText(N_L_LEFT_DLG_KEY)
            label_n_l_right.setText(N_L_RIGHT_DLG_KEY)

    def _update_fields_for_experiment(exp_n):
        is1 = exp_n == 1
        is2 = exp_n == 2
        # Exp1: tempo máx. procura + Salvar + HUD tempo total
        for w in (label_search, edit_search, check_save_search, check_search_hud):
            w.setVisible(is1)
        if is1:
            label_search.setText(search_max_s_dlg_key(padrao_search))
            check_save_search.setChecked(False)
        # Exp2: duração + Salvar + distratores
        for w in (
            dur_label,
            edit_dur,
            check_save_dur,
            label_n_l_left,
            edit_n_l_left,
            label_n_l_right,
            edit_n_l_right,
        ):
            w.setVisible(is2)
        if is2:
            dur_label.setText(exp2_duration_dlg_key(padrao_exp2))
            dval = float(default_exp2_duration_s)
            edit_dur.setText(str(int(dval) if dval == int(dval) else dval))
            check_save_dur.setChecked(False)
            _apply_session_n_l()
        dlg.adjustSize()

    def _on_experiment_changed(_idx=None):
        exp_n = 1 if combo_exp.currentIndex() == 0 else 2
        _fill_sessions(exp_n)
        _update_fields_for_experiment(exp_n)

    def _on_session_changed(_idx=None):
        if combo_exp.currentIndex() == 1:
            _apply_session_n_l()

    combo_exp.currentIndexChanged.connect(_on_experiment_changed)
    combo_session.currentIndexChanged.connect(_on_session_changed)

    _fill_sessions(de0)
    combo_session.blockSignals(True)
    idx0 = combo_session.findText(str(ds0))
    if idx0 >= 0:
        combo_session.setCurrentIndex(idx0)
    combo_session.blockSignals(False)
    _update_fields_for_experiment(de0)

    if dlg.exec_() != QtWidgets.QDialog.Accepted:
        return None

    p = (edit_participant.text() or "").strip()
    if not p:
        print("Código do participante em falta.", file=sys.stderr)
        return None
    exp_n = 1 if combo_exp.currentIndex() == 0 else 2
    try:
        s = int(combo_session.currentText().strip())
    except ValueError:
        print("Sessão inválida.", file=sys.stderr)
        return None
    smax_ok = session_condition_max_for_experiment(exp_n)
    if not (SESSION_CONDITION_MIN <= s <= smax_ok):
        print(
            "Sessão deve estar entre 1 e {0}.".format(smax_ok),
            file=sys.stderr,
        )
        return None
    try:
        r = int((edit_run.text() or "1").strip())
    except ValueError:
        print("Repetição inválida (inteiro >= 1).", file=sys.stderr)
        return None
    if r < 1:
        print("Repetição deve ser >= 1.", file=sys.stderr)
        return None
    wh = parse_wxh_string(str(combo_res.currentText() or "").strip())
    if wh is None:
        print(
            'Resolução inválida. Use formato Largura x Altura (ex.: “1680x1050”).',
            file=sys.stderr,
        )
        return None
    win_res_w, win_res_h = wh

    search_max_s = float(DEFAULT_SEARCH_MAX_S)
    if exp_n == 1:
        try:
            search_max_s = _validate_search_max_s(
                float(str(edit_search.text()).strip().replace(",", "."))
            )
        except ValueError:
            print(
                "Tempo máx. de procura inválido (número > 0, em segundos).",
                file=sys.stderr,
            )
            return None
        if check_save_search.isChecked():
            try:
                saved = save_search_max_s_padrao(search_max_s)
                print(
                    'Novo padrão tempo máx. procura gravado: {0:g} s → {1}'.format(
                        saved, _PREFS_FILE
                    )
                )
            except (OSError, ValueError) as exc:
                print(
                    "Aviso: não foi possível gravar o padrão de procura ({0}).".format(
                        exc
                    ),
                    file=sys.stderr,
                )
    try:
        cod_grey_ms = _validate_cod_grey_ms(
            float(str(edit_cod.text()).strip().replace(",", "."))
        )
    except ValueError:
        print(
            "Change over delay inválido (número >= 0, em milissegundos).",
            file=sys.stderr,
        )
        return None
    show_gaze_dot = bool(check_gaze.isChecked())
    show_search_time_hud = bool(check_search_hud.isChecked()) if exp_n == 1 else False

    exp2_duration_s = float(DEFAULT_EXP2_DURATION_S)
    n_L_left = None  # type: Optional[int]
    n_L_right = None  # type: Optional[int]
    if exp_n == 2:
        try:
            exp2_duration_s = _validate_exp2_duration_s(
                float(str(edit_dur.text()).strip().replace(",", "."))
            )
        except ValueError:
            print(
                "Tempo total Exp. 2 inválido (número > 0, em segundos; conta após instruções).",
                file=sys.stderr,
            )
            return None
        if check_save_dur.isChecked():
            try:
                saved = save_exp2_duration_padrao(exp2_duration_s)
                print(
                    'Novo padrão Exp. 2 gravado: {0:g} s (após instruções) → {1}'.format(
                        saved, _PREFS_FILE
                    )
                )
            except (OSError, ValueError) as exc:
                print(
                    "Aviso: não foi possível gravar o padrão Exp. 2 ({0}).".format(exc),
                    file=sys.stderr,
                )
        try:
            n_L_left = _validate_n_l_count(
                float(str(edit_n_l_left.text()).strip().replace(",", "."))
            )
            n_L_right = _validate_n_l_count(
                float(str(edit_n_l_right.text()).strip().replace(",", "."))
            )
        except ValueError:
            print(
                "Distratores L inválidos (inteiros >= 0 para esquerda e direita).",
                file=sys.stderr,
            )
            return None

    return (
        p,
        s,
        r,
        exp_n,
        win_res_w,
        win_res_h,
        search_max_s,
        cod_grey_ms,
        show_gaze_dot,
        show_search_time_hud,
        exp2_duration_s,
        n_L_left,
        n_L_right,
    )


def ask_experiment_setup_interactive(
    default_participant: str,
    default_session: int,
    default_run: int,
    default_experiment: int = 1,
    default_search_max_s: Optional[float] = None,
    default_cod_grey_ms: Optional[float] = None,
    default_show_gaze_dot: Optional[bool] = None,
    default_exp2_duration_s: Optional[float] = None,
    default_show_search_time_hud: Optional[bool] = None,
) -> Optional[
    Tuple[
        str,
        int,
        int,
        int,
        int,
        int,
        float,
        float,
        bool,
        bool,
        float,
        Optional[int],
        Optional[int],
    ]
]:
    """
    Diálogo antes da janela fullscreen.
    Em ambos: participante, experimento, sessão, repetição, COD, ponto do olhar, resolução.
    Só Exp1: tempo máx. de procura do T (+ Salvar) e HUD de tempo total de procura.
    Só Exp2: tempo total (+ Salvar) e distratores L esquerda/direita (atualizam com a sessão).
    Qt: mudar Experimento 1↔2 mostra/oculta esses campos no mesmo diálogo.
    Fallback DlgFromDict: se abrir como Exp1 e escolher Exp2, pede tempo/L num diálogo extra.
    """
    dr = max(1, int(default_run))
    de = max(EXPERIMENT_MIN, min(EXPERIMENT_MAX, int(default_experiment)))
    def_sms = (
        float(default_search_max_s)
        if default_search_max_s is not None
        else _default_search_max_s_from_env()
    )
    def_cod = (
        float(default_cod_grey_ms)
        if default_cod_grey_ms is not None
        else _default_cod_grey_ms_from_env()
    )
    def_gaze = (
        bool(default_show_gaze_dot)
        if default_show_gaze_dot is not None
        else _default_show_gaze_dot_from_env()
    )
    def_search_hud = (
        bool(default_show_search_time_hud)
        if default_show_search_time_hud is not None
        else _default_show_search_time_hud_from_env()
    )
    def_exp2_dur = (
        float(default_exp2_duration_s)
        if default_exp2_duration_s is not None
        else _default_exp2_duration_s_from_env_or_prefs()
    )

    qt_result = _ask_experiment_setup_qt(
        default_participant=default_participant,
        default_session=default_session,
        default_run=dr,
        default_experiment=de,
        default_search_max_s=def_sms,
        default_cod_grey_ms=def_cod,
        default_show_gaze_dot=def_gaze,
        default_exp2_duration_s=def_exp2_dur,
        default_show_search_time_hud=def_search_hud,
    )
    if qt_result is not None:
        return qt_result
    # Se Qt cancelou (None) vs Qt indisponível: _ask_experiment_setup_qt devolve None
    # tanto em cancel como se Qt falhar ao importar. Distinguir: se Qt existe e user
    # cancelou, não cair no fallback. Se Qt não existe, fallback.
    if _import_qt_widgets() is not None:
        return None

    try:
        from psychopy import gui
    except Exception:
        print(
            "Aviso: psychopy.gui indisponível. Defina --participant, --experiment, --session, "
            "--session-run ou FORR_PARTICIPANT_ID, FORR_EXPERIMENT (1|2), FORR_SESSION, "
            "FORR_SESSION_RUN (>=1).",
            file=sys.stderr,
        )
        return None

    smax = session_condition_max_for_experiment(de)
    sess_label = (
        "Sessão (condição 1 a {0})".format(smax)
        if smax != SESSION_CONDITION_MAX
        else "Sessão (condição 1 a 5)"
    )
    res_label = "Resolução da janela"
    env_res = parse_wxh_string(os.environ.get("FORR_RESOLUTION", "").strip())
    def_w = env_res[0] if env_res else DEFAULT_WINDOW_W
    def_h = env_res[1] if env_res else DEFAULT_WINDOW_H
    res_list = _resolution_dropdown_choices(def_w, def_h)
    padrao_exp2 = get_exp2_duration_padrao()
    exp2_dur_key = exp2_duration_dlg_key(padrao_exp2)
    padrao_search = get_search_max_s_padrao()
    search_max_key = search_max_s_dlg_key(padrao_search)
    # Campos específicos: procura máx. só Exp1; duração/L só Exp2.
    # Sessão + repetição entram sempre (registo + condições de cada experimento).
    show_exp1_fields = de == 1
    show_exp2_fields = de == 2
    sess_for_nl = max(
        SESSION_CONDITION_MIN,
        min(session_condition_max_for_experiment(2), int(default_session)),
    )
    shown_n_l_left, shown_n_l_right, shown_ratio = load_exp2_condition_n_l(sess_for_nl)
    n_l_left_key = N_L_LEFT_DLG_KEY
    n_l_right_key = N_L_RIGHT_DLG_KEY
    if shown_ratio:
        n_l_left_key = "Distratores L — esquerda ({0})".format(shown_ratio)
        n_l_right_key = "Distratores L — direita ({0})".format(shown_ratio)

    info = {
        "Código do participante": (default_participant or "").strip(),
        "Experimento": _experiment_choice_list(de),
        COD_GREY_MS_DLG_KEY: str(int(def_cod) if def_cod == int(def_cod) else def_cod),
        SHOW_GAZE_DOT_DLG_KEY: bool(def_gaze),
        sess_label: _session_choice_list(default_session, de),
        "Repetição desta condição (1=1ª vez, 2=2ª…)": str(dr),
        res_label: res_list,
    }
    order = [
        "Código do participante",
        "Experimento",
        sess_label,
        "Repetição desta condição (1=1ª vez, 2=2ª…)",
    ]
    if show_exp1_fields:
        info[search_max_key] = str(
            int(def_sms) if float(def_sms) == int(def_sms) else def_sms
        )
        info[SEARCH_MAX_SAVE_DLG_KEY] = False
        info[SHOW_SEARCH_TIME_HUD_DLG_KEY] = bool(def_search_hud)
        order.extend(
            [search_max_key, SEARCH_MAX_SAVE_DLG_KEY, SHOW_SEARCH_TIME_HUD_DLG_KEY]
        )
    if show_exp2_fields:
        info[exp2_dur_key] = str(
            int(def_exp2_dur) if def_exp2_dur == int(def_exp2_dur) else def_exp2_dur
        )
        info[EXP2_DURATION_SAVE_DLG_KEY] = False
        info[n_l_left_key] = str(shown_n_l_left)
        info[n_l_right_key] = str(shown_n_l_right)
        order.extend(
            [
                exp2_dur_key,
                EXP2_DURATION_SAVE_DLG_KEY,
                n_l_left_key,
                n_l_right_key,
            ]
        )
    order.extend(
        [
            COD_GREY_MS_DLG_KEY,
            SHOW_GAZE_DOT_DLG_KEY,
            res_label,
        ]
    )
    dlg = gui.DlgFromDict(
        dictionary=info,
        title="Forrageamento visual",
        order=order,
    )
    if not dlg.OK:
        return None
    p = (info.get("Código do participante") or "").strip()
    if not p:
        print("Código do participante em falta.", file=sys.stderr)
        return None
    exp_n = _parse_experiment_field(info.get("Experimento"))
    sess_raw = info.get(sess_label)
    if isinstance(sess_raw, list):
        sess_raw = sess_raw[0] if sess_raw else "1"
    try:
        s = int(str(sess_raw).strip())
    except ValueError:
        print("Sessão inválida.", file=sys.stderr)
        return None
    smax_ok = session_condition_max_for_experiment(exp_n)
    if not (SESSION_CONDITION_MIN <= s <= smax_ok):
        print(
            "Sessão deve estar entre 1 e {0}.".format(smax_ok),
            file=sys.stderr,
        )
        return None
    rep_key = "Repetição desta condição (1=1ª vez, 2=2ª…)"
    rep_raw = info.get(rep_key) or "1"
    try:
        r = int(str(rep_raw).strip())
    except ValueError:
        print("Repetição inválida (inteiro >= 1).", file=sys.stderr)
        return None
    if r < 1:
        print("Repetição deve ser >= 1.", file=sys.stderr)
        return None
    res_raw = info.get(res_label)
    if isinstance(res_raw, list):
        res_raw = res_raw[0] if res_raw else None
    wh = parse_wxh_string(str(res_raw or "").strip())
    if wh is None:
        print(
            'Resolução inválida. Use formato Largura x Altura (ex.: “1680x1050”).',
            file=sys.stderr,
        )
        return None
    win_res_w, win_res_h = wh
    search_max_s = float(DEFAULT_SEARCH_MAX_S)
    if exp_n == 1 and show_exp1_fields and search_max_key in info:
        sms_raw = info.get(search_max_key) or str(padrao_search)
        try:
            search_max_s = _validate_search_max_s(
                float(str(sms_raw).strip().replace(",", "."))
            )
        except ValueError:
            print(
                "Tempo máx. de procura inválido (número > 0, em segundos).",
                file=sys.stderr,
            )
            return None
        if _parse_dlg_bool(info.get(SEARCH_MAX_SAVE_DLG_KEY), default=False):
            try:
                saved = save_search_max_s_padrao(search_max_s)
                print(
                    'Novo padrão tempo máx. procura gravado: {0:g} s → {1}'.format(
                        saved, _PREFS_FILE
                    )
                )
            except (OSError, ValueError) as exc:
                print(
                    "Aviso: não foi possível gravar o padrão de procura ({0}).".format(
                        exc
                    ),
                    file=sys.stderr,
                )
    elif exp_n == 1:
        # Exp1 escolhido sem campo no formulário (abriu como Exp2): usa padrão gravado.
        search_max_s = float(padrao_search)
    cod_raw = info.get(COD_GREY_MS_DLG_KEY) or str(DEFAULT_COD_GREY_MS)
    try:
        cod_grey_ms = _validate_cod_grey_ms(float(str(cod_raw).strip().replace(",", ".")))
    except ValueError:
        print(
            "Change over delay inválido (número >= 0, em milissegundos).",
            file=sys.stderr,
        )
        return None
    show_gaze_dot = _parse_dlg_bool(
        info.get(SHOW_GAZE_DOT_DLG_KEY), default=DEFAULT_SHOW_GAZE_DOT
    )
    show_search_time_hud = False
    if exp_n == 1:
        if show_exp1_fields and SHOW_SEARCH_TIME_HUD_DLG_KEY in info:
            show_search_time_hud = _parse_dlg_bool(
                info.get(SHOW_SEARCH_TIME_HUD_DLG_KEY),
                default=DEFAULT_SHOW_SEARCH_TIME_HUD,
            )
        else:
            show_search_time_hud = bool(def_search_hud)

    # Tempo total + Salvar: exclusivos do Experimento 2.
    exp2_duration_s = float(DEFAULT_EXP2_DURATION_S)
    if exp_n == 2:
        if show_exp2_fields and exp2_dur_key in info:
            exp2_dur_raw = info.get(exp2_dur_key) or str(padrao_exp2)
            try:
                exp2_duration_s = _validate_exp2_duration_s(
                    float(str(exp2_dur_raw).strip().replace(",", "."))
                )
            except ValueError:
                print(
                    "Tempo total Exp. 2 inválido (número > 0, em segundos; conta após instruções).",
                    file=sys.stderr,
                )
                return None
            if _parse_dlg_bool(info.get(EXP2_DURATION_SAVE_DLG_KEY), default=False):
                try:
                    saved = save_exp2_duration_padrao(exp2_duration_s)
                    print(
                        'Novo padrão Exp. 2 gravado: {0:g} s (após instruções) → {1}'.format(
                            saved, _PREFS_FILE
                        )
                    )
                except (OSError, ValueError) as exc:
                    print(
                        "Aviso: não foi possível gravar o padrão Exp. 2 ({0}).".format(exc),
                        file=sys.stderr,
                    )
        else:
            # Diálogo principal abriu como Exp1; pedir tempo/L só do Exp2.
            follow = _ask_exp2_only_params_interactive(s, float(padrao_exp2))
            if follow is None:
                return None
            exp2_duration_s, n_L_left, n_L_right = follow
            return (
                p,
                s,
                r,
                exp_n,
                win_res_w,
                win_res_h,
                search_max_s,
                cod_grey_ms,
                show_gaze_dot,
                show_search_time_hud,
                exp2_duration_s,
                n_L_left,
                n_L_right,
            )

    n_L_left: Optional[int] = None
    n_L_right: Optional[int] = None
    if exp_n == 2:
        csv_l, csv_r, _ = load_exp2_condition_n_l(s)
        if show_exp2_fields and n_l_left_key in info:
            left_raw = info.get(n_l_left_key)
            right_raw = info.get(n_l_right_key)
            try:
                entered_l = _validate_n_l_count(
                    float(
                        str(left_raw if left_raw is not None else shown_n_l_left)
                        .strip()
                        .replace(",", ".")
                    )
                )
                entered_r = _validate_n_l_count(
                    float(
                        str(right_raw if right_raw is not None else shown_n_l_right)
                        .strip()
                        .replace(",", ".")
                    )
                )
            except ValueError:
                print(
                    "Distratores L inválidos (inteiros >= 0 para esquerda e direita).",
                    file=sys.stderr,
                )
                return None
            # Se mudou a sessão e não editou os L, usar os da nova condição.
            if (
                int(s) != int(sess_for_nl)
                and entered_l == shown_n_l_left
                and entered_r == shown_n_l_right
            ):
                n_L_left, n_L_right = csv_l, csv_r
            else:
                n_L_left, n_L_right = entered_l, entered_r
        else:
            n_L_left, n_L_right = csv_l, csv_r

    return (
        p,
        s,
        r,
        exp_n,
        win_res_w,
        win_res_h,
        search_max_s,
        cod_grey_ms,
        show_gaze_dot,
        show_search_time_hud,
        exp2_duration_s,
        n_L_left,
        n_L_right,
    )


def resolve_experiment_metadata(
    args: Any,
) -> Optional[
    Tuple[
        str,
        int,
        int,
        int,
        Optional[int],
        Optional[int],
        float,
        float,
        bool,
        bool,
        float,
        Optional[int],
        Optional[int],
    ]
]:
    """
    Participante + experimento (1|2) + sessão + repetição via CLI/env ou diálogo.
    Devolve (
        código participante, session_condition, session_run, experiment_number,
        gui_largura_ou_None, gui_altura_ou_None, search_max_s (Exp1), cod_grey_ms,
        show_gaze_dot, show_search_time_hud, exp2_duration_s, n_L_left, n_L_right
    ).
    Os índices de resolução são None quando os metadados vêm só por CLI/env (use resolve_window_size).
    n_L_* só aplicados no Exp2 (None no Exp1).
    """
    p_cli = (getattr(args, "participant", None) or "").strip()
    p_env = os.environ.get("FORR_PARTICIPANT_ID", "").strip()
    p = p_cli or p_env

    e_cli = getattr(args, "experiment", None)
    e_env = os.environ.get("FORR_EXPERIMENT", "").strip()
    e: Optional[int] = None
    if e_cli is not None:
        try:
            e = int(e_cli)
        except ValueError:
            print("--experiment deve ser 1 ou 2.", file=sys.stderr)
            return None
    elif e_env:
        try:
            e = int(e_env)
        except ValueError:
            print("FORR_EXPERIMENT deve ser 1 ou 2.", file=sys.stderr)
            return None

    s_cli = getattr(args, "session", None)
    r_cli = getattr(args, "session_run", None)
    s_env = os.environ.get("FORR_SESSION", "").strip()
    r_env = os.environ.get("FORR_SESSION_RUN", "").strip()

    s: Optional[int] = None
    r: Optional[int] = None
    if s_cli is not None:
        try:
            s = int(s_cli)
        except ValueError:
            print("--session deve ser inteiro (1–5 Exp1, 1–10 Exp2).", file=sys.stderr)
            return None
    elif s_env:
        try:
            s = int(s_env)
        except ValueError:
            print("FORR_SESSION deve ser inteiro.", file=sys.stderr)
            return None

    if r_cli is not None:
        try:
            r = int(r_cli)
        except ValueError:
            print("--session-run deve ser inteiro >= 1.", file=sys.stderr)
            return None
    elif r_env:
        try:
            r = int(r_env)
        except ValueError:
            print("FORR_SESSION_RUN deve ser inteiro >= 1.", file=sys.stderr)
            return None

    if p and s is not None and r is not None:
        if e is None:
            e = 1
        if e < EXPERIMENT_MIN or e > EXPERIMENT_MAX:
            print("Experimento deve ser 1 ou 2.", file=sys.stderr)
            return None
        smax = session_condition_max_for_experiment(e)
        if not (SESSION_CONDITION_MIN <= s <= smax):
            print(
                "Sessão (condição) deve estar entre 1 e {0}.".format(smax),
                file=sys.stderr,
            )
            return None
        if r < 1:
            print("Repetição da sessão deve ser >= 1.", file=sys.stderr)
            return None
        try:
            search_max_s = resolve_search_max_seconds(args, e)
            cod_grey_ms = resolve_cod_grey_ms(args)
            show_gaze_dot = resolve_show_gaze_dot(args)
            show_search_time_hud = (
                resolve_show_search_time_hud(args) if e == 1 else False
            )
            exp2_duration_s = resolve_exp2_duration_s(args, e)
            n_L_left, n_L_right = resolve_exp2_n_l(args, s, e)
        except ValueError:
            return None
        return (
            p,
            s,
            r,
            e,
            None,
            None,
            search_max_s,
            cod_grey_ms,
            show_gaze_dot,
            show_search_time_hud,
            exp2_duration_s,
            n_L_left,
            n_L_right,
        )

    ds = s if s is not None else 1
    dr = r if r is not None else 1
    de = e if e is not None else 1
    try:
        def_sms = resolve_search_max_seconds(args, de)
        def_cod = resolve_cod_grey_ms(args)
        def_gaze = resolve_show_gaze_dot(args)
        def_search_hud = resolve_show_search_time_hud(args)
        def_exp2_dur = resolve_exp2_duration_s(args, de)
    except ValueError:
        return None
    return ask_experiment_setup_interactive(
        default_participant=p or "",
        default_session=ds,
        default_run=dr,
        default_experiment=de,
        default_search_max_s=def_sms,
        default_cod_grey_ms=def_cod,
        default_show_gaze_dot=def_gaze,
        default_exp2_duration_s=def_exp2_dur,
        default_show_search_time_hud=def_search_hud,
    )


def resolve_conditions_path(args: Any, experiment_number: int) -> str:
    """Ficheiro de condições por defeito conforme o experimento, salvo --conditions."""
    if getattr(args, "conditions", None):
        return str(args.conditions)
    if experiment_number == 1:
        return _CONDITIONS_FILE_EXP1
    return _CONDITIONS_FILE_EXP2_SESSIONS


def get_gaze_xy_pix(tracker: Any, win: visual.Window) -> Optional[Tuple[float, float]]:
    if tracker is None:
        return None
    sample = tracker.getNewestSample()
    if sample is None:
        return None
    eye = None
    if sample.isRightSample() and sample.getRightEye() is not None:
        eye = sample.getRightEye()
    elif sample.isLeftSample() and sample.getLeftEye() is not None:
        eye = sample.getLeftEye()
    if eye is None:
        return None
    gaze = eye.getGaze()
    if gaze is None:
        return None
    if len(gaze) < 2:
        return None
    if not math.isfinite(float(gaze[0])) or not math.isfinite(float(gaze[1])):
        return None
    # EyeLink usa valores muito grandes (por exemplo MISSING_DATA) durante
    # piscadas/perda de amostra. Não os transforme num cursor visível.
    if (
        float(gaze[0]) < 0.0
        or float(gaze[0]) > float(win.size[0])
        or float(gaze[1]) < 0.0
        or float(gaze[1]) > float(win.size[1])
    ):
        return None
    gx = gaze[0] - win.size[0] / 2.0
    gy = (win.size[1] / 2.0) - gaze[1]
    return gx, gy


def read_gaze_xy(
    dummy_mode: bool,
    mouse: event.Mouse,
    tracker: Any,
    win: visual.Window,
) -> Optional[Tuple[float, float]]:
    """Amostra mais recente do olhar (EyeLink) ou posição do rato (dummy)."""
    if dummy_mode:
        return tuple(mouse.getPos())
    return get_gaze_xy_pix(tracker, win)


def in_circle(
    x: float, y: float, center: Tuple[float, float], radius: float
) -> bool:
    cx, cy = center
    return ((x - cx) ** 2 + (y - cy) ** 2) <= radius**2


def in_rect(
    gx: float,
    gy: float,
    cx: float,
    cy: float,
    half_w: float,
    half_h: float,
) -> bool:
    return abs(gx - cx) <= half_w and abs(gy - cy) <= half_h


def half_left_bounds(sw: float, sh: float) -> Tuple[float, float, float, float]:
    """Returns cx, cy, half_w, half_h for left monitor half AOI."""
    return -sw / 4.0, 0.0, sw / 4.0, sh / 2.0


def half_right_bounds(sw: float, sh: float) -> Tuple[float, float, float, float]:
    return sw / 4.0, 0.0, sw / 4.0, sh / 2.0


def region_choice_neutral_half_w(sw: float) -> float:
    """Metade da largura da zona neutra (|x| ≤ isto não escolhe lado)."""
    return max(REGION_CHOICE_NEUTRAL_HALF_W_MIN_PX, sw * REGION_CHOICE_NEUTRAL_HALF_W_FRAC)


def psychopy_aabb_to_eyelink_screen_px(
    cx: float,
    cy: float,
    half_w: float,
    half_h: float,
    sw: float,
    sh: float,
) -> Tuple[int, int, int, int]:
    """
    Caixa alinhada aos eixos: coords PsychoPy (origem centro, +y para cima) → pixeis de
    ecrã EyeLink (origem canto superior esquerdo, +y para baixo), coerente com
    get_gaze_xy_pix/gaze[0], gaze[1] e DISPLAY_COORDS em setup_eyelink.
    Devolve (left, top, right, bottom) inclusive.
    """
    left_c = cx - half_w
    right_c = cx + half_w
    top_c = cy + half_h
    bottom_c = cy - half_h
    l = int(round(left_c + sw / 2.0))
    r = int(round(right_c + sw / 2.0))
    t = int(round(sh / 2.0 - top_c))
    b = int(round(sh / 2.0 - bottom_c))
    sw_i = max(1, int(sw))
    sh_i = max(1, int(sh))
    l = max(0, min(sw_i - 1, l))
    r = max(0, min(sw_i - 1, r))
    t = max(0, min(sh_i - 1, t))
    b = max(0, min(sh_i - 1, b))
    if l > r:
        l, r = r, l
    if t > b:
        t, b = b, t
    return l, t, r, b


def send_dataviewer_ia_rect(
    tracker: Any,
    ia_id: int,
    label: str,
    left: int,
    top: int,
    right: int,
    bottom: int,
) -> None:
    """
    Define retângulo de Interest Area para importação no EyeLink Data Viewer.
    Formato SR: !V IAREA RECTANGLE id left top right bottom label
    (label sem espaços para compatibilidade).
    """
    if tracker is None:
        return
    lab = "".join((c if (c.isalnum() or c in "_-") else "_") for c in label)[:48]
    msg = "!V IAREA RECTANGLE {0} {1} {2} {3} {4} {5}".format(
        ia_id, left, top, right, bottom, lab or "IA_{0}".format(ia_id)
    )
    send_msg(tracker, msg)


def emit_dataviewer_ias_drift(
    tracker: Any, drift_r: float, sw: float, sh: float
) -> None:
    """Bounding box do círculo de fixação central (mesmo raio que in_circle no drift)."""
    if tracker is None:
        return
    r = float(drift_r)
    l, t, rpx, b = psychopy_aabb_to_eyelink_screen_px(0.0, 0.0, r, r, sw, sh)
    send_msg(
        tracker,
        "IA_META drift use in_circle center=0,0 radius_px={0} bbox_ltrb={1},{2},{3},{4}".format(
            int(drift_r), l, t, rpx, b
        ),
    )
    send_dataviewer_ia_rect(tracker, 101, "drift_cross_bbox", l, t, rpx, b)


def emit_dataviewer_ias_region_choice(
    tracker: Any,
    neutral_hw: float,
    sw: float,
    sh: float,
) -> None:
    """IAs da escolha Esquerda/Direita: neutro visual + bandas gx<-nh e gx>nh (|gy|<=sh/2)."""
    if tracker is None:
        return
    nh = float(neutral_hw)
    half_h_n = (sh * 0.5) / 2.0
    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(0.0, 0.0, nh, half_h_n, sw, sh)
    send_msg(tracker, "IA_BLOCK region_choice neutral_visual white_frame")
    send_dataviewer_ia_rect(tracker, 102, "region_neutral", l, t, r, b)

    hw_band = (sw / 2.0 - nh) / 2.0
    cx_left_band = (-sw / 2.0 + (-nh)) / 2.0
    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(
        cx_left_band, 0.0, hw_band, sh / 2.0, sw, sh
    )
    send_msg(tracker, "IA_META region_left logic gx_lt_neg_neutral_hw")
    send_dataviewer_ia_rect(tracker, 103, "region_left", l, t, r, b)

    cx_right_band = (nh + sw / 2.0) / 2.0
    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(
        cx_right_band, 0.0, hw_band, sh / 2.0, sw, sh
    )
    send_msg(tracker, "IA_META region_right logic gx_gt_neutral_hw")
    send_dataviewer_ia_rect(tracker, 104, "region_right", l, t, r, b)


def emit_dataviewer_ias_forage_panels(
    tracker: Any,
    sw: float,
    sh: float,
    panel_w: float,
    panel_h: float,
) -> None:
    """Metades COD + painéis L/T (AOIs em run_trial)."""
    if tracker is None:
        return
    hpw = panel_w / 2.0
    hph = panel_h / 2.0
    send_msg(tracker, "IA_BLOCK foraging halves_and_panels")

    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(
        -sw / 4.0, 0.0, sw / 4.0, sh / 2.0, sw, sh
    )
    send_dataviewer_ia_rect(tracker, 105, "forage_half_left", l, t, r, b)

    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(
        sw / 4.0, 0.0, sw / 4.0, sh / 2.0, sw, sh
    )
    send_dataviewer_ia_rect(tracker, 106, "forage_half_right", l, t, r, b)

    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(-sw / 4.0, 0.0, hpw, hph, sw, sh)
    send_dataviewer_ia_rect(tracker, 107, "forage_panel_left", l, t, r, b)

    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(sw / 4.0, 0.0, hpw, hph, sw, sh)
    send_dataviewer_ia_rect(tracker, 108, "forage_panel_right", l, t, r, b)


def emit_dataviewer_ia_target_t(
    tracker: Any,
    sw: float,
    sh: float,
    cx: float,
    cy: float,
    half_w: float,
    half_h: float,
) -> None:
    """Interest Area em volta do T (confirmação: olhar dentro + Espaço)."""
    if tracker is None:
        return
    l, t, r, b = psychopy_aabb_to_eyelink_screen_px(cx, cy, half_w, half_h, sw, sh)
    send_msg(
        tracker,
        "IA_META target_T_hit psychopy_cx_cy_halfwh={0:.1f},{1:.1f},{2:.1f},{3:.1f}".format(
            cx, cy, half_w, half_h
        ),
    )
    send_msg(tracker, "IA_BLOCK target_T_confirm gaze_in_rect_plus_space")
    send_dataviewer_ia_rect(tracker, 110, "target_T_hit", l, t, r, b)


def emit_dataviewer_ias_exp2_dual_targets(
    tracker: Any,
    sw: float,
    sh: float,
    t_aoi_left,
    t_aoi_right,
) -> None:
    """Exp2: duas IAs nos T (110 = esquerda, 111 = direita); confirmação olhar + Espaço em qualquer um."""
    if tracker is None:
        return
    send_msg(
        tracker,
        "IA_META exp2_dual_T psychopy IAREA 110=left 111=right",
    )
    send_msg(tracker, "IA_BLOCK exp2_dual_T_confirm gaze_in_rect_plus_space")
    if t_aoi_left is not None:
        tcx, tcy, thw, thh = t_aoi_left
        l, t, r, b = psychopy_aabb_to_eyelink_screen_px(tcx, tcy, thw, thh, sw, sh)
        send_dataviewer_ia_rect(tracker, 110, "target_T_left", l, t, r, b)
    if t_aoi_right is not None:
        tcx, tcy, thw, thh = t_aoi_right
        l, t, r, b = psychopy_aabb_to_eyelink_screen_px(tcx, tcy, thw, thh, sw, sh)
        send_dataviewer_ia_rect(tracker, 111, "target_T_right", l, t, r, b)


def emit_dataviewer_ias_discrete_fullscreen(tracker: Any, sw: float, sh: float) -> None:
    """Área de análise da fase de resposta discreta (ecrã completo)."""
    if tracker is None:
        return
    sw_i = max(1, int(sw))
    sh_i = max(1, int(sh))
    send_msg(tracker, "IA_META discrete full_screen RT arrows")
    send_dataviewer_ia_rect(tracker, 109, "discrete_fullscreen", 0, 0, sw_i - 1, sh_i - 1)


def make_region_choice_stims(
    win: visual.Window, sw: float, sh: float, sc: float
) -> List[Any]:
    neutral_hw = region_choice_neutral_half_w(sw)
    # Faixa neutra vertical: metade da altura do ecrã; marcos só por linhas à esquerda e à direita.
    neutral_rect_h = sh * 0.5
    outline_w = max(1, int(1.5 * sc))
    hl = max(44.0, 58.0 * sc)
    # Setas dentro da área ativa (fora da faixa neutra), aproximando-se do contorno.
    arrow_inset = max(neutral_hw + hl * 0.35, sw * 0.18)
    left_x = min(-arrow_inset, -neutral_hw - hl * 0.25)
    right_x = max(arrow_inset, neutral_hw + hl * 0.25)
    neutral_fill = visual.Rect(
        win,
        width=2.0 * neutral_hw,
        height=neutral_rect_h,
        pos=(0.0, 0.0),
        fillColor=BACKGROUND_COLOR,
        lineColor=None,
        units="pix",
    )
    line_thick = float(max(2.0, outline_w))
    neutral_left_line = visual.Rect(
        win,
        width=line_thick,
        height=neutral_rect_h,
        pos=(-neutral_hw + line_thick / 2.0, 0.0),
        fillColor=(1.0, 1.0, 1.0),
        lineColor=None,
        units="pix",
    )
    neutral_right_line = visual.Rect(
        win,
        width=line_thick,
        height=neutral_rect_h,
        pos=(neutral_hw - line_thick / 2.0, 0.0),
        fillColor=(1.0, 1.0, 1.0),
        lineColor=None,
        units="pix",
    )
    left = visual.TextStim(
        win,
        text="◄",
        pos=(left_x, 0.0),
        height=hl,
        color=(0.55, 0.55, 0.55),
        units="pix",
    )
    right = visual.TextStim(
        win,
        text="►",
        pos=(right_x, 0.0),
        height=hl,
        color=(0.55, 0.55, 0.55),
        units="pix",
    )
    # Texto só dentro da faixa neutra (largura ≤ zona |x|≤neutral_hw) para ler sem
    # activar escolha esq./dir.; centrado verticalmente na banda (altura sh*0.5 em y=0).
    cue_wrap = max(120.0, 2.0 * neutral_hw * 0.88)
    cue_h = max(15.0, min(22.0 * sc, neutral_rect_h * 0.14))
    cue = visual.TextStim(
        win,
        text="Olhe para a esquerda ou para a direita para iniciar.",
        pos=(0.0, 0.0),
        height=cue_h,
        color=(1.0, 1.0, 1.0),
        units="pix",
        wrapWidth=cue_wrap,
        font=INSTRUCTIONS_UI_FONT,
        bold=True,
    )
    # Fundo da faixa neutra + limite só por duas linhas verticais (sem traço em cima/baixo).
    return [neutral_fill, neutral_left_line, neutral_right_line, left, right, cue]


def fit_size_contain(nat_w: float, nat_h: float, box_w: float, box_h: float) -> Tuple[float, float]:
    """Escala mantendo proporção para caber na caixa."""
    if nat_w <= 0 or nat_h <= 0 or box_w <= 0 or box_h <= 0:
        return (box_w, box_h)
    s = min(box_w / float(nat_w), box_h / float(nat_h))
    return (max(1.0, nat_w * s), max(1.0, nat_h * s))


def layout_scale(sw: float, sh: float) -> float:
    """Factor ~1 no 1280×768; maior em monitores maiores."""
    return min(float(sw) / REF_LAYOUT_W, float(sh) / REF_LAYOUT_H)


def procedural_panel_size(box_w: float, box_h: float) -> Tuple[float, float]:
    """Mesma proporção em ambos os lados; encaixa na metade útil do ecrã."""
    return fit_size_contain(
        float(PANEL_NAT_FALLBACK[0]),
        float(PANEL_NAT_FALLBACK[1]),
        box_w,
        box_h,
    )


def search_panel_grid_dims(panel_w: float, panel_h: float) -> Tuple[int, int]:
    """Dimensões da grelha lógica igual à de build_search_panel_stims (densidade)."""
    _density = min(panel_w, panel_h) / 10.0
    cell_base = max(36.0, min(56.0, _density))
    n_cols = max(5, int(panel_w / cell_base))
    n_rows = max(6, int(panel_h / (cell_base * 1.05)))
    return n_cols, n_rows


def t_position_trial_id(
    cx: float,
    cy: float,
    tx: float,
    ty: float,
    panel_w: float,
    panel_h: float,
    n_cols: int,
    n_rows: int,
) -> int:
    """
    ID inteiro 1…(n_cols×n_rows) da célula (grelha sobre o painel, canto sup-esquerdo do painel)
    onde cai o centro do T; mesmo id = mesma região espacial relativa ao painel alvo (para comparar
    participantes com a mesma resolução de janela e painel).
    """
    cell_w = panel_w / float(n_cols)
    cell_h = panel_h / float(n_rows)
    left = cx - panel_w / 2.0
    top = cy + panel_h / 2.0
    lx = tx - left
    ly_from_top = top - ty
    if cell_w <= 1e-9:
        cell_w = 1.0
    if cell_h <= 1e-9:
        cell_h = 1.0
    col = int(max(0, min(n_cols - 1, math.floor(lx / cell_w))))
    row = int(max(0, min(n_rows - 1, math.floor(ly_from_top / cell_h))))
    return row * n_cols + col + 1


def make_fixation_cross(
    win: visual.Window, arm_px: float, bar_px: float, color: Tuple[float, float, float]
) -> List[visual.Rect]:
    h = visual.Rect(
        win,
        width=arm_px * 2,
        height=bar_px,
        pos=(0, 0),
        fillColor=color,
        lineColor=None,
        units="pix",
    )
    v = visual.Rect(
        win,
        width=bar_px,
        height=arm_px * 2,
        pos=(0, 0),
        fillColor=color,
        lineColor=None,
        units="pix",
    )
    return [h, v]


def make_gaze_dot(win: visual.Window, scale: float = 1.0) -> visual.Circle:
    """Marcador neutro e pequeno; área idêntica à esquerda e à direita."""
    radius = max(3.0, GAZE_DOT_RADIUS_PX * float(scale))
    return visual.Circle(
        win,
        radius=radius,
        edges=16,
        pos=(0.0, 0.0),
        fillColor=GAZE_DOT_FILL,
        lineColor=GAZE_DOT_LINE,
        lineWidth=max(1.0, 1.5 * float(scale)),
        units="pix",
    )


def draw_gaze_dot(
    gaze_dot: visual.Circle,
    gaze_xy: Optional[Tuple[float, float]],
    show: bool = True,
) -> None:
    """Desenha somente quando show=True e há uma amostra válida do olhar/rato."""
    if not show or gaze_xy is None:
        return
    gaze_dot.pos = gaze_xy
    gaze_dot.draw()


def draw_gaze_dot_fresh(
    gaze_dot: visual.Circle,
    dummy_mode: bool,
    mouse: event.Mouse,
    tracker: Any,
    win: visual.Window,
    show: bool = True,
) -> None:
    """
    Reamostra o olhar imediatamente antes de desenhar o ponto.
    Reduz atraso visual vs amostrar no início do frame (antes de desenhar muitos L).
    """
    if not show:
        return
    draw_gaze_dot(gaze_dot, read_gaze_xy(dummy_mode, mouse, tracker, win), show=True)


def _scatter_positions_in_rect(
    cx: float,
    cy: float,
    panel_w: float,
    panel_h: float,
    n_points: int,
    min_dist: float,
    margin: float,
    rng: random.Random,
) -> List[Tuple[float, float]]:
    """
    Posições pseudo-aleatórias com distância mínima entre centros (rejeição),
    para evitar alinhamentos tipo texto/grelha.
    """
    half_w = panel_w / 2.0 - margin
    half_h = panel_h / 2.0 - margin
    if half_w <= 1.0 or half_h <= 1.0 or n_points <= 0:
        return []
    md = min_dist
    best: List[Tuple[float, float]] = []
    for _relax in range(18):
        out: List[Tuple[float, float]] = []
        min_d2 = md * md
        max_attempts = max(12000, n_points * 250)
        attempts = 0
        while len(out) < n_points and attempts < max_attempts:
            attempts += 1
            x = cx + rng.uniform(-half_w, half_w)
            y = cy + rng.uniform(-half_h, half_h)
            ok = True
            for ox, oy in out:
                dx = x - ox
                dy = y - oy
                if dx * dx + dy * dy < min_d2:
                    ok = False
                    break
            if ok:
                out.append((x, y))
        if len(out) >= n_points:
            return out[:n_points]
        if len(out) > len(best):
            best = out
        md *= 0.9
    return best[:n_points] if len(best) >= n_points else best


def build_instruction_example_panel_stims(
    win: visual.Window,
    cx: float,
    cy: float,
    panel_w: float,
    panel_h: float,
    font: Optional[str] = None,
) -> List[Any]:
    """Caixa pequena nas instruções: L dispersas + um T verde (layout reprodutível)."""
    if font is None:
        font = INSTRUCTIONS_UI_FONT
    rng = random.Random(12345)
    stims: List[Any] = []
    stims.append(
        visual.Rect(
            win,
            width=panel_w + 6,
            height=panel_h + 6,
            pos=(cx, cy),
            fillColor=INSTRUCTIONS_PREVIEW_BOX_FILL,
            lineColor=INSTRUCTIONS_PREVIEW_BOX_LINE,
            lineWidth=2,
            units="pix",
        )
    )
    n_items = max(14, min(24, int(min(panel_w, panel_h) / 11.0)))
    letter_h = max(12.0, min(panel_w, panel_h) / 7.5)
    min_center_dist = letter_h * 0.82
    inset = max(letter_h * 0.35, 4.0)
    positions = _scatter_positions_in_rect(
        cx, cy, panel_w, panel_h, n_items, min_center_dist, inset, rng
    )
    if len(positions) < n_items:
        hw = max(panel_w / 2.0 - inset, 6.0)
        hh = max(panel_h / 2.0 - inset, 6.0)
        while len(positions) < n_items:
            positions.append(
                (cx + rng.uniform(-hw, hw), cy + rng.uniform(-hh, hh))
            )
    rng.shuffle(positions)
    t_index = rng.randrange(len(positions))
    for i, (x, y) in enumerate(positions):
        if i == t_index:
            ch = "T"
            col = INSTRUCTIONS_PREVIEW_TARGET_GREEN
        else:
            ch = "L"
            col = rng.choice(GRAYS_FOR_L)
        ori = float(rng.choice((0, 90, 180, 270)))
        stims.append(
            visual.TextStim(
                win,
                text=ch,
                pos=(x, y),
                height=letter_h,
                color=col,
                ori=ori,
                units="pix",
                bold=True,
                font=font,
            )
        )
    return stims


def build_search_panel_stims(
    win: visual.Window,
    cx: float,
    cy: float,
    panel_w: float,
    panel_h: float,
    include_target: bool,
    rng: random.Random,
) -> Tuple[
    List[Any],
    Optional[Tuple[float, float, float, float]],
    Optional[int],
]:
    """
    L e T dispersos no painel (distância mínima), orientações e cinzas variados.
    Segundo: (tx, ty, half_w, half_h) para AOI do T, ou None.
    Terceiro: trial_id da posição do T (célula da grelha 1…n_cols×n_rows), ou None no painel só-L.
    """
    stims: List[Any] = []
    border = visual.Rect(
        win,
        width=panel_w + 4,
        height=panel_h + 4,
        pos=(cx, cy),
        lineColor=(0.12, 0.12, 0.12),
        fillColor=None,
        lineWidth=2,
        units="pix",
    )
    stims.append(border)

    _density = min(panel_w, panel_h) / 10.0
    cell_base = max(36.0, min(56.0, _density))
    n_cols = max(5, int(panel_w / cell_base))
    n_rows = max(6, int(panel_h / (cell_base * 1.05)))
    cell_w = panel_w / float(n_cols)
    cell_h = panel_h / float(n_rows)
    letter_h = min(cell_w, cell_h) * 0.68
    n_grid = n_cols * n_rows
    n_items = max(
        SEARCH_PANEL_STIM_COUNT_MIN,
        int(n_grid * SEARCH_PANEL_STIM_COUNT_FRAC),
    )
    min_center_dist = letter_h * 0.82
    inset = max(letter_h * 0.38, 6.0)
    positions = _scatter_positions_in_rect(
        cx, cy, panel_w, panel_h, n_items, min_center_dist, inset, rng
    )
    if not positions:
        hw = max(panel_w / 2.0 - inset, 8.0)
        hh = max(panel_h / 2.0 - inset, 8.0)
        for _ in range(min(n_items, 48)):
            positions.append(
                (cx + rng.uniform(-hw, hw), cy + rng.uniform(-hh, hh))
            )
    if not positions:
        return stims, None, None
    rng.shuffle(positions)
    t_index = rng.randrange(len(positions)) if include_target else -1

    target_aoi: Optional[Tuple[float, float, float, float]] = None
    t_trial_id: Optional[int] = None
    hw_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_W_FRAC)
    hh_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_H_FRAC)

    for i, (x, y) in enumerate(positions):
        if include_target and i == t_index:
            ch = "T"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
            target_aoi = (float(x), float(y), float(hw_t), float(hh_t))
            t_trial_id = t_position_trial_id(
                cx, cy, float(x), float(y), panel_w, panel_h, n_cols, n_rows
            )
        else:
            ch = "L"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
        stims.append(
            visual.TextStim(
                win,
                text=ch,
                pos=(x, y),
                height=letter_h,
                color=col_rgb,
                ori=ori,
                units="pix",
                bold=True,
            )
        )
    return stims, target_aoi, t_trial_id


def build_search_panel_stims_n_l(
    win: visual.Window,
    cx: float,
    cy: float,
    panel_w: float,
    panel_h: float,
    n_l_distractors: int,
    include_target: bool,
    rng: random.Random,
):
    """
    Painel com exactamente n_l_distractors letras L e opcionalmente um T (alvo).
    Usado no Exp2 (contagens iniciais e após decaimento/repor).
    """
    stims: List[Any] = []
    border = visual.Rect(
        win,
        width=panel_w + 4,
        height=panel_h + 4,
        pos=(cx, cy),
        lineColor=(0.12, 0.12, 0.12),
        fillColor=None,
        lineWidth=2,
        units="pix",
    )
    stims.append(border)

    _density = min(panel_w, panel_h) / 10.0
    cell_base = max(36.0, min(56.0, _density))
    n_cols = max(5, int(panel_w / cell_base))
    n_rows = max(6, int(panel_h / (cell_base * 1.05)))
    cell_w = panel_w / float(n_cols)
    cell_h = panel_h / float(n_rows)
    letter_h = min(cell_w, cell_h) * 0.68
    n_l_distractors = max(0, int(n_l_distractors))
    n_items = n_l_distractors + (1 if include_target else 0)
    if n_items < 1:
        return stims, None, None

    min_center_dist = letter_h * 0.82
    inset = max(letter_h * 0.38, 6.0)
    positions = _scatter_positions_in_rect(
        cx, cy, panel_w, panel_h, n_items, min_center_dist, inset, rng
    )
    if not positions:
        hw = max(panel_w / 2.0 - inset, 8.0)
        hh = max(panel_h / 2.0 - inset, 8.0)
        for _ in range(min(n_items, 96)):
            positions.append(
                (cx + rng.uniform(-hw, hw), cy + rng.uniform(-hh, hh))
            )
    if not positions:
        return stims, None, None
    rng.shuffle(positions)
    t_index = rng.randrange(len(positions)) if include_target else -1

    target_aoi: Optional[Tuple[float, float, float, float]] = None
    t_trial_id: Optional[int] = None
    hw_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_W_FRAC)
    hh_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_H_FRAC)

    for i, (x, y) in enumerate(positions):
        if include_target and i == t_index:
            ch = "T"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
            target_aoi = (float(x), float(y), float(hw_t), float(hh_t))
            t_trial_id = t_position_trial_id(
                cx, cy, float(x), float(y), panel_w, panel_h, n_cols, n_rows
            )
        else:
            ch = "L"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
        stims.append(
            visual.TextStim(
                win,
                text=ch,
                pos=(x, y),
                height=letter_h,
                color=col_rgb,
                ori=ori,
                units="pix",
                bold=True,
            )
        )
    return stims, target_aoi, t_trial_id


def build_exp2_fixed_lt_panel(
    win: visual.Window,
    cx: float,
    cy: float,
    panel_w: float,
    panel_h: float,
    n_l_distractors: int,
    include_target: bool,
    rng: random.Random,
):
    """
    Exp2: grelha de posições fixas (sem reembaralhar após o scatter). Os L só devem
    ser ocultados depois; o T mantém-se no mesmo sítio. Devolve (all_stims, l_stims, t_stim, target_aoi).
    """
    border = visual.Rect(
        win,
        width=panel_w + 4,
        height=panel_h + 4,
        pos=(cx, cy),
        lineColor=(0.12, 0.12, 0.12),
        fillColor=None,
        lineWidth=2,
        units="pix",
    )
    _density = min(panel_w, panel_h) / 10.0
    cell_base = max(36.0, min(56.0, _density))
    n_cols = max(5, int(panel_w / cell_base))
    n_rows = max(6, int(panel_h / (cell_base * 1.05)))
    cell_w = panel_w / float(n_cols)
    cell_h = panel_h / float(n_rows)
    letter_h = min(cell_w, cell_h) * 0.68
    n_l_distractors = max(0, int(n_l_distractors))
    n_slots = n_l_distractors + (1 if include_target else 0)
    if n_slots < 1:
        return [border], [], None, None

    min_center_dist = letter_h * 0.82
    inset = max(letter_h * 0.38, 6.0)
    positions = _scatter_positions_in_rect(
        cx, cy, panel_w, panel_h, n_slots, min_center_dist, inset, rng
    )
    if not positions:
        hw = max(panel_w / 2.0 - inset, 8.0)
        hh = max(panel_h / 2.0 - inset, 8.0)
        for _ in range(min(n_slots, 96)):
            positions.append(
                (cx + rng.uniform(-hw, hw), cy + rng.uniform(-hh, hh))
            )
    if not positions:
        return [border], [], None, None
    while len(positions) < n_slots:
        hw = max(panel_w / 2.0 - inset, 8.0)
        hh = max(panel_h / 2.0 - inset, 8.0)
        positions.append(
            (cx + rng.uniform(-hw, hw), cy + rng.uniform(-hh, hh))
        )

    # Não embaralhar: ordem das posições fixa para esta semente.
    t_slot = rng.randrange(n_slots) if include_target else -1
    hw_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_W_FRAC)
    hh_t = max(TARGET_T_AOI_MIN_HALF_PX, letter_h * TARGET_T_AOI_HALF_H_FRAC)
    target_aoi: Optional[Tuple[float, float, float, float]] = None
    l_stims: List[Any] = []
    letter_stims: List[Any] = []
    t_stim = None

    for i in range(n_slots):
        x, y = positions[i]
        if include_target and i == t_slot:
            ch = "T"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
            target_aoi = (float(x), float(y), float(hw_t), float(hh_t))
            ts = visual.TextStim(
                win,
                text=ch,
                pos=(x, y),
                height=letter_h,
                color=col_rgb,
                ori=ori,
                units="pix",
                bold=True,
            )
            t_stim = ts
            letter_stims.append(ts)
        else:
            ch = "L"
            ori = float(rng.choice((0, 90, 180, 270)))
            col_rgb = rng.choice(GRAYS_FOR_L)
            ls = visual.TextStim(
                win,
                text=ch,
                pos=(x, y),
                height=letter_h,
                color=col_rgb,
                ori=ori,
                units="pix",
                bold=True,
            )
            try:
                ls.opacity = 1.0
            except Exception:
                pass
            l_stims.append(ls)
            letter_stims.append(ls)

    all_stims: List[Any] = [border] + letter_stims
    return all_stims, l_stims, t_stim, target_aoi


def make_arrow_prompt_stims(win: visual.Window, sw: float, sh: float, sc: float) -> List[Any]:
    """Setas grandes para a resposta final (só desenho; teclado igual)."""
    hl = max(36.0, 48.0 * sc)
    return [
        visual.TextStim(
            win,
            text="◄ Esquerda",
            pos=(-sw * 0.28, -sh * 0.08),
            height=hl * 0.55,
            color=(0.45, 0.45, 0.45),
            units="pix",
        ),
        visual.TextStim(
            win,
            text="Direita ►",
            pos=(sw * 0.28, -sh * 0.08),
            height=hl * 0.55,
            color=(0.45, 0.45, 0.45),
            units="pix",
        ),
    ]


def make_hourglass_stim(
    win: visual.Window,
    pos: Tuple[float, float],
    sc: float,
    fill_color: Optional[Tuple[float, float, float]] = None,
    line_color: Optional[Tuple[float, float, float]] = None,
) -> Any:
    """
    Marcador no lado inativo: quadrado cinza (nome histórico `hourglass_stim`; sem PNG).
    Por defeito mais escuro que BACKGROUND_COLOR; pode passar fill_color/line_color (ex.: Exp2).
    """
    fill = SIDE_MARKER_FILL if fill_color is None else fill_color
    line = SIDE_MARKER_LINE if line_color is None else line_color
    side = max(36.0, 52.0 * sc)
    lw = max(1.0, 2.0 * sc)
    return visual.Rect(
        win,
        width=side,
        height=side,
        pos=pos,
        fillColor=fill,
        lineColor=line,
        lineWidth=lw,
        units="pix",
    )


def draw_stim_batch(stims: List[Any]) -> None:
    for s in stims:
        s.draw()


def setup_eyelink(win: visual.Window, edf_name: str, tracker_address: str) -> Any:
    tracker = pylink.EyeLink(tracker_address)
    tracker.openDataFile(edf_name)
    tracker.setOfflineMode()
    pylink.pumpDelay(100)

    scn_width, scn_height = win.size
    tracker.sendCommand(f"screen_pixel_coords = 0 0 {scn_width - 1} {scn_height - 1}")
    tracker.sendMessage(f"DISPLAY_COORDS 0 0 {scn_width - 1} {scn_height - 1}")
    tracker.sendMessage(f"VIEWING_DISTANCE_CM {int(VIEWING_DISTANCE_CM)}")

    tracker.sendCommand(
        "file_event_filter = LEFT,RIGHT,FIXATION,SACCADE,BLINK,MESSAGE,BUTTON,INPUT"
    )
    tracker.sendCommand(
        "file_sample_data = LEFT,RIGHT,GAZE,HREF,RAW,AREA,GAZERES,STATUS,INPUT"
    )
    tracker.sendCommand(
        "link_event_filter = LEFT,RIGHT,FIXATION,FIXUPDATE,SACCADE,BLINK,BUTTON,INPUT"
    )
    tracker.sendCommand(
        "link_sample_data = LEFT,RIGHT,GAZE,GAZERES,AREA,STATUS,INPUT"
    )

    genv = EyeLinkCoreGraphicsPsychoPy(tracker, win)
    pylink.openGraphicsEx(genv)
    tracker.doTrackerSetup()

    tracker.setOfflineMode()
    pylink.pumpDelay(100)
    tracker.startRecording(1, 1, 1, 1)
    pylink.pumpDelay(100)
    tracker.sendMessage("SYNCTIME")
    tracker.sendMessage(
        "IA_SESSION_META forrageamento_exp1 DISPLAY_COORDS_AND samples_match_psychopy_gaze "
        "VIEWING_DISTANCE_CM {0}".format(int(VIEWING_DISTANCE_CM))
    )
    return tracker


def shutdown_eyelink(tracker: Any, edf_name: str, local_folder: str) -> None:
    if tracker is None:
        return
    os.makedirs(local_folder, exist_ok=True)
    local_path = os.path.join(local_folder, edf_name)
    try:
        tracker.stopRecording()
    except Exception:
        pass
    try:
        tracker.setOfflineMode()
        pylink.pumpDelay(100)
    except Exception:
        pass
    try:
        tracker.closeDataFile()
        pylink.pumpDelay(500)
    except Exception:
        pass
    try:
        tracker.receiveDataFile(edf_name, local_path)
        print(f"EDF saved to: {local_path}")
    except Exception as e:
        print(f"EDF transfer failed: {e}")
    try:
        tracker.close()
    except Exception:
        pass


def run_eyelink_recalibration_runtime(
    win: visual.Window, tracker: Any, tag: str = "runtime"
) -> None:
    """
    Pausa gravação, abre setup/calibração EyeLink com graphics PsychoPy e retoma SYNCTIME.
    Semelhante ao fluxo em Conditional Discrimination… (tecla X).
    """
    if tracker is None:
        return
    tag_safe = str(tag or "runtime").replace(" ", "_")
    send_msg(tracker, "EYELINK_RECALIB_START {0}".format(tag_safe))
    try:
        tracker.stopRecording()
    except Exception:
        pass
    try:
        tracker.setOfflineMode()
        pylink.pumpDelay(100)
    except Exception:
        pass

    genv = EyeLinkCoreGraphicsPsychoPy(tracker, win)
    try:
        try:
            pylink.openGraphicsEx(genv)
        except Exception as e:
            send_msg(tracker, "EYELINK_RECALIB_OPEN_GFX_FAIL {0}".format(e))
            raise
        try:
            tracker.doTrackerSetup()
        finally:
            try:
                pylink.closeGraphics()
            except Exception:
                pass
        tracker.setOfflineMode()
        pylink.pumpDelay(100)
        tracker.startRecording(1, 1, 1, 1)
        pylink.pumpDelay(100)
        tracker.sendMessage("SYNCTIME")
        send_msg(tracker, "EYELINK_RECALIB_END {0}".format(tag_safe))
    except Exception as e:
        print("Recalibração EyeLink falhou: {0}".format(e), file=sys.stderr)
        send_msg(tracker, "EYELINK_RECALIB_FAILED {0} err={1}".format(tag_safe, e))
        try:
            tracker.setOfflineMode()
            pylink.pumpDelay(80)
            tracker.startRecording(1, 1, 1, 1)
            pylink.pumpDelay(80)
            tracker.sendMessage("SYNCTIME")
            send_msg(tracker, "EYELINK_RECORDING_RECOVER_AFTER_RECALIB_FAIL")
        except Exception as e2:
            send_msg(tracker, "EYELINK_RECORDING_RECOVER_FAIL {0}".format(e2))


def show_recalibration_pause_screen(
    win: visual.Window,
    kb,
    tracker: Any,
) -> bool:
    """
    Ecrã de pausa antes do setup EyeLink. Retorna False se T+Esc pediu saída (QUIT_SAVE_REQUESTED).
    Durante o loop não aceita novos atalhos X/C (evita recursão).
    """
    sw, sh = float(win.size[0]), float(win.size[1])
    msg = visual.TextStim(
        win,
        text=(
            "Experimento em pausa.\n"
            "Pressione ESPAÇO para abrir a recalibração / validação do EyeLink.\n"
            "Tecla X ou C — novo pedido de pausa e recalibração durante a tarefa.\n"
            "(T e depois Esc — gravar e sair.)"
        ),
        pos=(0.0, 0.0),
        height=max(18.0, min(26.0, sh * 0.024)),
        color=(1.0, 1.0, 1.0),
        units="pix",
        wrapWidth=min(sw * 0.82, 920.0),
        bold=True,
        font=INSTRUCTIONS_UI_FONT,
    )
    flush_keyboard(kb)
    while True:
        poll_global_quit_t_esc(kb, tracker, win, allow_recalibration_hotkeys=False)
        if QUIT_SAVE_REQUESTED:
            return False
        keys = kb.getKeys(keyList=["space"], waitRelease=False)
        if keys:
            break
        try:
            ev_keys = event.getKeys(keyList=["space"])
        except TypeError:
            ev_keys = event.getKeys()
        if "space" in ev_keys:
            break
        win.color = BACKGROUND_COLOR
        msg.draw()
        win.flip()
    flush_keyboard(kb)
    return True


def _consume_recalibration_hotkey(kb, key_char: str) -> bool:
    """True se a tecla foi premida (legacy Keyboard ou event.getKeys)."""
    kl = key_char.lower()
    variants = [kl, kl.upper()]
    got = False
    try:
        for ev in kb.getKeys(keyList=variants, waitRelease=False):
            if getattr(ev, "name", "").lower() == kl:
                got = True
    except Exception:
        pass
    try:
        for k in event.getKeys(keyList=variants):
            if str(k).lower() == kl:
                got = True
    except TypeError:
        for k in event.getKeys():
            if str(k).lower() == kl:
                got = True
                break
    return got


class TrialLog(object):
    """Registo por trial (sem dataclasses: compatível com Python 3.6)."""

    def __init__(self, trial_index=0):
        self.trial_index = trial_index
        self.correctkey = ""
        self.alvos = ""
        self.fundoesquerda = ""
        self.fundodireita = ""
        self.target_side = ""
        self.active_side_after_choice = ""
        # Tempos com estímulo visível na fase de forrageamento (exclui intervalo COD sem painéis)
        self.forage_time_left_s = 0.0
        self.forage_time_right_s = 0.0
        self.cod_switch_count = 0
        self.discrete_key = ""
        self.accuracy = ""
        self.rt_discrete_s = -1.0
        self.target_found = False
        self.search_timed_out = False
        # Segundos desde olhar centrado (drift_ok) até achar o T ou search_timeout; -1 se N/A
        self.search_time_s = -1.0
        self.aborted = False
        self.trial_id = ""
        self.target_t_x = None  # type: Optional[float]
        self.target_t_y = None  # type: Optional[float]
        self.extra = {}
        # Nível segmento (forrageamento): preenchido em run_trial
        self.segment_rows = []  # type: List[Dict[str, Any]]
        self.event_rows = []  # type: List[Dict[str, Any]]
        self._trial_clock = None  # type: Any
        self._foraging_clock = None  # type: Any
        self._forage_seg_start_t = None  # type: Optional[float]
        self._forage_seg_side = None  # type: Optional[str]


def _trial_append_event(log: TrialLog, event_type: str, side: str = "") -> None:
    tc = log._trial_clock
    if tc is None:
        return
    log.event_rows.append(
        {
            "event_index": len(log.event_rows),
            "event_type": event_type,
            "side": side,
            "t_trial_s": round(float(tc.getTime()), 6),
        }
    )


def _forage_open_segment(log: TrialLog, foraging_clock: Any, side: str) -> None:
    log._forage_seg_start_t = float(foraging_clock.getTime())
    log._forage_seg_side = side


def _forage_close_segment(
    log: TrialLog,
    foraging_clock: Any,
    switched_to: str,
    found_target_during_segment: bool,
) -> None:
    if log._forage_seg_start_t is None or log._forage_seg_side is None:
        return
    end_t = float(foraging_clock.getTime())
    start_t = float(log._forage_seg_start_t)
    dur = end_t - start_t
    log.segment_rows.append(
        {
            "segment_index": len(log.segment_rows),
            "side": log._forage_seg_side,
            "segment_start_s": round(start_t, 6),
            "segment_end_s": round(end_t, 6),
            "duration_s": round(dur, 6),
            "switched_to": switched_to,
            "found_target_during_segment": 1 if found_target_during_segment else 0,
        }
    )
    log._forage_seg_start_t = None
    log._forage_seg_side = None


def send_msg(tracker: Any, text: str) -> None:
    if tracker is not None:
        try:
            tracker.sendMessage(text)
        except Exception:
            pass
    print(f"MSG: {text}")


def reset_session_quit_flags() -> None:
    global _t_esc_arm_until, QUIT_SAVE_REQUESTED
    _t_esc_arm_until = 0.0
    QUIT_SAVE_REQUESTED = False


def poll_global_quit_t_esc(
    kb,
    tracker: Any = None,
    win: Any = None,
    allow_recalibration_hotkeys: bool = True,
) -> None:
    """T+Esc encerra com gravação; X ou C pausam e permitem recalibração EyeLink (com ecrã + Espaço)."""
    global _t_esc_arm_until, QUIT_SAVE_REQUESTED
    if QUIT_SAVE_REQUESTED:
        return
    now = core.getTime()
    if _t_esc_arm_until > 0.0 and now > _t_esc_arm_until:
        _t_esc_arm_until = 0.0

    if allow_recalibration_hotkeys and tracker is not None and win is not None:
        got_x = _consume_recalibration_hotkey(kb, "x")
        got_c = _consume_recalibration_hotkey(kb, "c")
        hotkey = None
        if got_x:
            hotkey = "key_X"
        elif got_c:
            hotkey = "key_C"
        if hotkey is not None:
            send_msg(tracker, "RECALIBRATION_REQUEST {0}".format(hotkey))
            _t_esc_arm_until = 0.0
            flush_keyboard(kb)
            try:
                if show_recalibration_pause_screen(win, kb, tracker):
                    run_eyelink_recalibration_runtime(win, tracker, tag=hotkey)
                send_msg(tracker, "RECALIBRATION_RESUME {0}".format(hotkey))
            except Exception as e:
                print("Fluxo de recalibração: {0}".format(e), file=sys.stderr)
                send_msg(tracker, "RECALIBRATION_FLOW_ERROR {0} {1}".format(hotkey, e))
            flush_keyboard(kb)
            return

    got_t = False
    for ev in kb.getKeys(keyList=["t"], waitRelease=False):
        if getattr(ev, "name", "") == "t":
            got_t = True
    try:
        if "t" in event.getKeys(keyList=["t"]):
            got_t = True
    except TypeError:
        for k in event.getKeys():
            if str(k).lower() == "t":
                got_t = True
                break
    if got_t:
        _t_esc_arm_until = now + T_ESC_ARM_TIMEOUT_S

    if _t_esc_arm_until > now:
        got_esc = False
        for ev in kb.getKeys(keyList=["escape"], waitRelease=False):
            if getattr(ev, "name", "") == "escape":
                got_esc = True
        try:
            if "escape" in event.getKeys(keyList=["escape"]):
                got_esc = True
        except TypeError:
            for k in event.getKeys():
                if str(k).lower() == "escape":
                    got_esc = True
                    break
        if got_esc:
            QUIT_SAVE_REQUESTED = True
            _t_esc_arm_until = 0.0
            send_msg(tracker, "SESSION_QUIT T_ESC")


def run_drift_and_region_choice(
    win: visual.Window,
    kb,
    mouse: event.Mouse,
    tracker: Any,
    dummy_mode: bool,
    show_gaze_dot: bool = DEFAULT_SHOW_GAZE_DOT,
) -> Optional[str]:
    """
    Fixação central + escolha de metade (mesma lógica que o início de run_trial).
    Retorna 'left' ou 'right', ou None se QUIT_SAVE_REQUESTED.
    """
    sw, sh = win.size[0], win.size[1]
    sc = layout_scale(sw, sh)
    drift_r = max(24, int(DRIFT_AOI_RADIUS * sc))

    send_msg(tracker, "PHASE drift_start")
    emit_dataviewer_ias_drift(tracker, float(drift_r), float(sw), float(sh))
    fix_arm = max(14.0, 22.0 * sc)
    fix_bar = max(2.0, 4.0 * sc)
    drift_cross = make_fixation_cross(win, fix_arm, fix_bar, FIXATION_GRAY)
    gaze_dot = make_gaze_dot(win, sc)
    drift_clock = core.Clock()
    fix_entry = None
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            return None
        if dummy_mode:
            gaze_xy = tuple(mouse.getPos())
        else:
            gaze_xy = get_gaze_xy_pix(tracker, win)
        gx, gy = gaze_xy if gaze_xy is not None else (float("inf"), float("inf"))
        inside = in_circle(gx, gy, (0.0, 0.0), float(drift_r))
        now = drift_clock.getTime()
        if inside:
            if fix_entry is None:
                fix_entry = now
            elif (now - fix_entry) * 1000.0 >= DRIFT_FIX_MS:
                break
        else:
            fix_entry = None
        draw_stim_batch(drift_cross)
        draw_gaze_dot_fresh(
            gaze_dot, dummy_mode, mouse, tracker, win, show=show_gaze_dot
        )
        win.flip()
    send_msg(tracker, "PHASE drift_ok")

    send_msg(tracker, "PHASE region_choice_start")
    choice_clock = core.Clock()
    dwell_entry: Optional[float] = None
    dwell_side: Optional[str] = None
    active_side: str = "left"

    region_choice_stims = make_region_choice_stims(win, float(sw), float(sh), sc)
    neutral_hw = region_choice_neutral_half_w(float(sw))
    send_msg(
        tracker,
        "REGION_CHOICE neutral_half_w_px {:.1f} (faixa central sem escolha)".format(neutral_hw),
    )
    emit_dataviewer_ias_region_choice(tracker, float(neutral_hw), float(sw), float(sh))
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            return None
        if dummy_mode:
            gaze_xy = tuple(mouse.getPos())
        else:
            gaze_xy = get_gaze_xy_pix(tracker, win)
        gx, gy = gaze_xy if gaze_xy is not None else (float("inf"), float("inf"))
        in_band = abs(gy) <= float(sh) / 2.0
        in_neutral = in_band and abs(gx) <= neutral_hw
        in_l = in_band and gx < -neutral_hw
        in_r = in_band and gx > neutral_hw
        elapsed_ms = choice_clock.getTime() * 1000.0
        now_t = choice_clock.getTime()

        if elapsed_ms >= PRE_CHOICE_HOLD_MS:
            if in_neutral or not in_band:
                dwell_entry = None
                dwell_side = None
            elif in_l or in_r:
                side = "left" if in_l else "right"
                if dwell_side != side:
                    dwell_side = side
                    dwell_entry = now_t
                elif dwell_entry is not None:
                    if (now_t - dwell_entry) * 1000.0 >= REGION_CHOICE_MIN_DWELL_MS:
                        active_side = side
                        break
            else:
                dwell_entry = None
                dwell_side = None

        win.color = BACKGROUND_COLOR
        draw_stim_batch(region_choice_stims)
        draw_gaze_dot_fresh(
            gaze_dot, dummy_mode, mouse, tracker, win, show=show_gaze_dot
        )
        win.flip()
    send_msg(tracker, "PHASE region_chosen {0}".format(active_side))
    return active_side


def show_discrete_response_feedback(
    win: visual.Window,
    kb,
    tracker: Any,
    is_correct: bool,
    duration_s: float = CORRECT_FEEDBACK_S,
) -> bool:
    """
    Após a escolha discreta (←/→): ecrã cheio verde (correto) ou cinza (errado) durante duration_s.
    Devolve False se T+Esc durante o feedback.
    """
    phase = "correct_feedback" if is_correct else "incorrect_feedback"
    fb_color = CORRECT_FEEDBACK_COLOR if is_correct else BACKGROUND_COLOR
    send_msg(tracker, "PHASE {0}".format(phase))
    win.color = fb_color
    win.flip()
    fb_clock = core.Clock()
    while fb_clock.getTime() < float(duration_s):
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            win.color = BACKGROUND_COLOR
            win.flip()
            return False
        win.color = fb_color
        win.flip()
    win.color = BACKGROUND_COLOR
    win.flip()
    return True


def run_trial(
    win: visual.Window,
    kb,
    mouse: event.Mouse,
    tracker: Any,
    trial: Dict[str, str],
    trial_index: int,
    dummy_mode: bool,
    stim_layout_seed: int,
    search_max_s: float = DEFAULT_SEARCH_MAX_S,
    cod_grey_ms: float = DEFAULT_COD_GREY_MS,
    show_gaze_dot: bool = DEFAULT_SHOW_GAZE_DOT,
    show_search_time_hud: bool = DEFAULT_SHOW_SEARCH_TIME_HUD,
    cumulative_search_time_s: float = 0.0,
) -> TrialLog:
    sw, sh = win.size[0], win.size[1]
    sc = layout_scale(sw, sh)
    drift_r = max(24, int(DRIFT_AOI_RADIUS * sc))
    margin_x = max(8, int(LAYOUT_MARGIN_X * sc))
    margin_y = max(8, int(LAYOUT_MARGIN_Y * sc))
    log = TrialLog(trial_index=trial_index)
    log.correctkey = trial.get("correctkey", "")
    log.alvos = trial.get("alvos", "")
    log.fundoesquerda = trial.get("fundoesquerda", "")
    log.fundodireita = trial.get("fundodireita", "")

    target_side = "left" if log.correctkey.strip().lower() == "left" else "right"
    log.target_side = target_side

    trial_clock = core.Clock()
    log._trial_clock = trial_clock
    _trial_append_event(log, "trial_start", "")

    send_msg(tracker, "TRIALID {0}".format(trial_index))
    send_msg(
        tracker,
        "CONDITION_TRIAL_ROW {0}".format((trial.get("trial") or "").strip()),
    )
    send_msg(tracker, "STIM_LAYOUT_SEED {0}".format(int(stim_layout_seed)))
    send_msg(tracker, "TARGET_SIDE {0}".format(target_side))
    send_msg(tracker, "SEARCH_MAX_S {0:.4f}".format(float(search_max_s)))
    send_msg(tracker, "COD_GREY_MS {0:.1f}".format(float(cod_grey_ms)))
    send_msg(
        tracker,
        "SHOW_GAZE_DOT {0}".format(1 if show_gaze_dot else 0),
    )
    send_msg(
        tracker,
        "SHOW_SEARCH_TIME_HUD {0}".format(1 if show_search_time_hud else 0),
    )

    # --- Drift / central fixation ---
    send_msg(tracker, "PHASE drift_start")
    _trial_append_event(log, "drift_start", "")
    emit_dataviewer_ias_drift(tracker, float(drift_r), float(sw), float(sh))
    fix_arm = max(14.0, 22.0 * sc)
    fix_bar = max(2.0, 4.0 * sc)
    drift_cross = make_fixation_cross(win, fix_arm, fix_bar, FIXATION_GRAY)
    gaze_dot = make_gaze_dot(win, sc)
    drift_clock = core.Clock()
    fix_entry = None
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            log.aborted = True
            log.extra["reason"] = "quit_t_esc"
            _trial_append_event(log, "quit_t_esc", "")
            _trial_append_event(log, "trial_abort", "")
            return log
        if dummy_mode:
            gaze_xy = tuple(mouse.getPos())
        else:
            gaze_xy = get_gaze_xy_pix(tracker, win)
        gx, gy = gaze_xy if gaze_xy is not None else (float("inf"), float("inf"))
        inside = in_circle(gx, gy, (0.0, 0.0), float(drift_r))
        now = drift_clock.getTime()
        if inside:
            if fix_entry is None:
                fix_entry = now
            elif (now - fix_entry) * 1000.0 >= DRIFT_FIX_MS:
                break
        else:
            fix_entry = None
        draw_stim_batch(drift_cross)
        draw_gaze_dot_fresh(
            gaze_dot, dummy_mode, mouse, tracker, win, show=show_gaze_dot
        )
        win.flip()
    send_msg(tracker, "PHASE drift_ok")
    _trial_append_event(log, "drift_ok", "")

    # Janela do limite de procura do T (search_max_s), por trial: arranca com o olhar já
    # centrado (drift_ok) e termina ao passar à pergunta lateral. Não conta as instruções
    # nem a fixação central, e não corre durante a resposta esquerda/direita.
    search_clock = core.Clock()
    send_msg(
        tracker,
        "SEARCH_WINDOW_START after_drift_ok limit_s {0:.4f}".format(float(search_max_s)),
    )

    # --- Region choice (gaze dwell 200 ms) after minimum display 800 ms ---
    send_msg(tracker, "PHASE region_choice_start")
    _trial_append_event(log, "region_choice_start", "")
    choice_clock = core.Clock()
    dwell_entry: Optional[float] = None
    dwell_side: Optional[str] = None
    active_side: str = "left"
    region_choice_timed_out = False

    region_choice_stims = make_region_choice_stims(win, float(sw), float(sh), sc)
    neutral_hw = region_choice_neutral_half_w(float(sw))
    send_msg(
        tracker,
        "REGION_CHOICE neutral_half_w_px {:.1f} (faixa central sem escolha)".format(neutral_hw),
    )
    emit_dataviewer_ias_region_choice(tracker, float(neutral_hw), float(sw), float(sh))
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            log.aborted = True
            log.extra["reason"] = "quit_t_esc"
            log.search_time_s = round(float(search_clock.getTime()), 6)
            _trial_append_event(log, "quit_t_esc", "")
            _trial_append_event(log, "trial_abort", "")
            return log
        if search_clock.getTime() >= float(search_max_s):
            region_choice_timed_out = True
            break
        if dummy_mode:
            gaze_xy = tuple(mouse.getPos())
        else:
            gaze_xy = get_gaze_xy_pix(tracker, win)
        gx, gy = gaze_xy if gaze_xy is not None else (float("inf"), float("inf"))
        in_band = abs(gy) <= float(sh) / 2.0
        in_neutral = in_band and abs(gx) <= neutral_hw
        in_l = in_band and gx < -neutral_hw
        in_r = in_band and gx > neutral_hw
        elapsed_ms = choice_clock.getTime() * 1000.0
        now_t = choice_clock.getTime()

        if elapsed_ms >= PRE_CHOICE_HOLD_MS:
            if in_neutral or not in_band:
                dwell_entry = None
                dwell_side = None
            elif in_l or in_r:
                side = "left" if in_l else "right"
                if dwell_side != side:
                    dwell_side = side
                    dwell_entry = now_t
                elif dwell_entry is not None:
                    if (now_t - dwell_entry) * 1000.0 >= REGION_CHOICE_MIN_DWELL_MS:
                        active_side = side
                        break
            else:
                dwell_entry = None
                dwell_side = None

        win.color = BACKGROUND_COLOR
        draw_stim_batch(region_choice_stims)
        draw_gaze_dot_fresh(
            gaze_dot, dummy_mode, mouse, tracker, win, show=show_gaze_dot
        )
        win.flip()
    if region_choice_timed_out:
        # Limite atingido sem escolher metade: sem lado registado; o ciclo de procura
        # abaixo deteta o mesmo limite e passa logo à pergunta lateral.
        send_msg(tracker, "REGION_CHOICE_TIMEOUT no_side_before_search_limit")
        _trial_append_event(log, "region_choice_timeout", "")
    else:
        log.active_side_after_choice = active_side
        send_msg(tracker, f"PHASE region_chosen {active_side}")
        _trial_append_event(log, "region_chosen", active_side)

    # --- Foraging + COD + target ---
    send_msg(tracker, "PHASE foraging_start")
    _trial_append_event(log, "foraging_start", "")
    foraging_clock = core.Clock()
    log._foraging_clock = foraging_clock
    _forage_open_segment(log, foraging_clock, active_side)
    _trial_append_event(log, "entry_{0}".format(active_side), active_side)
    cod_switch_count = 0
    cod_grey_until = 0.0
    in_cod_grey = False
    inactive_dwell_entry: Optional[float] = None
    time_left_s = 0.0
    time_right_s = 0.0
    last_forage_tick_t = foraging_clock.getTime()

    left_cx, left_cy = -sw / 4.0, 0.0
    right_cx, right_cy = sw / 4.0, 0.0
    half_inner_w = sw / 2.0 - 2.0 * margin_x
    max_panel_h = sh - 2.0 * margin_y
    panel_w, panel_h = procedural_panel_size(half_inner_w, max_panel_h)
    rng_l = random.Random((stim_layout_seed ^ 0xCAFEBABECAFEBABE) & 0x7FFFFFFFFFFFFFFF)
    rng_r = random.Random((stim_layout_seed ^ 0xA5A5A5A5A5A5A5A5) & 0x7FFFFFFFFFFFFFFF)
    stims_left, target_t_aoi_left, t_tid_left = build_search_panel_stims(
        win,
        left_cx,
        left_cy,
        panel_w,
        panel_h,
        target_side == "left",
        rng_l,
    )
    stims_right, target_t_aoi_right, t_tid_right = build_search_panel_stims(
        win,
        right_cx,
        right_cy,
        panel_w,
        panel_h,
        target_side == "right",
        rng_r,
    )
    target_t_aoi = (
        target_t_aoi_left if target_side == "left" else target_t_aoi_right
    )
    _tid = t_tid_left if target_side == "left" else t_tid_right
    log.trial_id = str(int(_tid)) if _tid is not None else ""
    log.target_t_x = None
    log.target_t_y = None
    if target_t_aoi is not None:
        log.target_t_x = round(float(target_t_aoi[0]), 4)
        log.target_t_y = round(float(target_t_aoi[1]), 4)
        send_msg(
            tracker,
            "TARGET_T_PSYCHOPY_XY {0} {1}".format(log.target_t_x, log.target_t_y),
        )
    _gnc, _gnr = search_panel_grid_dims(float(panel_w), float(panel_h))
    send_msg(
        tracker,
        "T_POSITION_TRIAL_ID {0} grid {1}x{2}".format(
            log.trial_id or "NA", _gnc, _gnr
        ),
    )
    emit_dataviewer_ias_forage_panels(
        tracker, float(sw), float(sh), float(panel_w), float(panel_h)
    )
    if target_t_aoi is not None:
        _ttx, _tty, _twh, _thh = target_t_aoi
        emit_dataviewer_ia_target_t(
            tracker, float(sw), float(sh), _ttx, _tty, _twh, _thh
        )
        tcx_t, tcy_t, thw_t, thh_t = _ttx, _tty, _twh, _thh
    else:
        tcx_t = left_cx if target_side == "left" else right_cx
        tcy_t = 0.0
        thw_t, thh_t = panel_w / 2.0, panel_h / 2.0
        send_msg(
            tracker,
            "IA_META target_T_hit FALLBACK panel_AOI (no T bbox — unexpected)",
        )
    hg_x = right_cx if active_side == "left" else left_cx
    hourglass: Any = make_hourglass_stim(
        win,
        (hg_x, 0.0),
        sc,
        fill_color=EXP2_INACTIVE_SIDE_FILL,
        line_color=EXP2_INACTIVE_SIDE_LINE,
    )

    target_fix_accum = 0.0
    target_found = False
    last_frame_t = core.getTime()

    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            log.aborted = True
            log.extra["reason"] = "quit_t_esc"
            log.cod_switch_count = cod_switch_count
            log.forage_time_left_s = time_left_s
            log.forage_time_right_s = time_right_s
            log.search_time_s = round(float(search_clock.getTime()), 6)
            _forage_close_segment(log, foraging_clock, "quit_t_esc", False)
            _trial_append_event(log, "quit_t_esc", "")
            _trial_append_event(log, "trial_abort", "")
            return log

        now = foraging_clock.getTime()
        search_elapsed = search_clock.getTime()
        if search_elapsed >= float(search_max_s):
            send_msg(tracker, "SEARCH_TIMEOUT elapsed_s {:.4f}".format(search_elapsed))
            _forage_close_segment(log, foraging_clock, "search_timeout", False)
            _trial_append_event(log, "search_timeout", "")
            log.search_timed_out = True
            log.search_time_s = round(float(search_elapsed), 6)
            break

        if in_cod_grey:
            if now >= cod_grey_until:
                in_cod_grey = False
                active_side = "right" if active_side == "left" else "left"
                inactive_dwell_entry = None
                last_forage_tick_t = now
                send_msg(tracker, f"COD_SWITCH side_active {active_side}")
                _trial_append_event(log, "cod_end", "")
                _forage_open_segment(log, foraging_clock, active_side)
                _trial_append_event(log, "entry_{0}".format(active_side), active_side)
            else:
                win.color = BACKGROUND_COLOR
                win.flip()
                continue

        # Tempo com array visível por lado (não conta o intervalo COD sem estímulos nem este frame até ao flip)
        dt_forage = now - last_forage_tick_t
        last_forage_tick_t = now
        if active_side == "left":
            time_left_s += dt_forage
        else:
            time_right_s += dt_forage

        if dummy_mode:
            gaze_xy = tuple(mouse.getPos())
        else:
            gaze_xy = get_gaze_xy_pix(tracker, win)
        gx, gy = gaze_xy if gaze_xy is not None else (float("inf"), float("inf"))

        inactive = "right" if active_side == "left" else "left"
        icx, icy, ihw, ihh = (
            half_left_bounds(sw, sh) if inactive == "left" else half_right_bounds(sw, sh)
        )
        in_inactive = in_rect(gx, gy, icx, icy, ihw, ihh)

        if in_inactive:
            if inactive_dwell_entry is None:
                inactive_dwell_entry = now
            elif (now - inactive_dwell_entry) * 1000.0 >= COD_FIX_MS:
                send_msg(tracker, "COD_START")
                nxt_side = "right" if active_side == "left" else "left"
                _forage_close_segment(log, foraging_clock, nxt_side, False)
                _trial_append_event(log, "cod_start", inactive)
                cod_switch_count += 1
                in_cod_grey = True
                cod_grey_until = now + float(cod_grey_ms) / 1000.0
                inactive_dwell_entry = None
                continue
        else:
            inactive_dwell_entry = None

        # AOI = rect em torno do glifo T (coerente com IA 110 no EDF); painel inteiro só em fallback.
        in_target_aoi = (
            active_side == target_side
            and in_rect(gx, gy, tcx_t, tcy_t, thw_t, thh_t)
        )
        keys = kb.getKeys(keyList=["space"], waitRelease=False)
        now_frame = core.getTime()
        dt = min(max(now_frame - last_frame_t, 0.0), 0.1)
        last_frame_t = now_frame
        if dt <= 0.0:
            dt = 1.0 / 60.0
        if in_target_aoi:
            target_fix_accum += dt
        else:
            target_fix_accum = 0.0

        if (
            active_side == target_side
            and in_target_aoi
            and target_fix_accum >= TARGET_FIX_MS / 1000.0
            and len(keys) > 0
        ):
            target_found = True
            log.target_found = True
            log.search_time_s = round(float(search_clock.getTime()), 6)
            send_msg(
                tracker,
                "TARGET_FOUND space_confirm gaze_in_target_T_IA search_time_s {:.4f}".format(
                    log.search_time_s
                ),
            )
            _forage_close_segment(log, foraging_clock, "target_found", True)
            _trial_append_event(log, "target_found", target_side)
            _trial_append_event(log, "response_space", "")
            break

        win.color = BACKGROUND_COLOR
        if active_side == "left":
            draw_stim_batch(stims_left)
            hourglass.pos = (right_cx, 0.0)
            hourglass.draw()
        else:
            draw_stim_batch(stims_right)
            hourglass.pos = (left_cx, 0.0)
            hourglass.draw()
        draw_gaze_dot_fresh(
            gaze_dot, dummy_mode, mouse, tracker, win, show=show_gaze_dot
        )
        win.flip()

    log.cod_switch_count = cod_switch_count
    log.forage_time_left_s = time_left_s
    log.forage_time_right_s = time_right_s
    send_msg(
        tracker,
        "FORAGE_TIMES_EXCL_COD left_s {:.4f} right_s {:.4f}".format(
            time_left_s, time_right_s
        ),
    )

    # --- Discrete choice ---
    send_msg(tracker, "PHASE discrete_choice_start")
    _trial_append_event(log, "discrete_choice_start", "")
    emit_dataviewer_ias_discrete_fullscreen(tracker, float(sw), float(sh))
    kb.clearEvents()
    prompt = visual.TextStim(
        win,
        text="Em que lado estava a letra T (alvo)?",
        pos=(0, sh * 0.22),
        height=max(18.0, int(28 * sc)),
        color=(0.65, 0.65, 0.65),
        units="pix",
        wrapWidth=sw * 0.88,
        font=INSTRUCTIONS_UI_FONT,
        bold=True,
    )
    arrow_hint_stims = make_arrow_prompt_stims(win, float(sw), float(sh), sc)
    search_hud_stim = None
    if show_search_time_hud:
        trial_search = (
            float(log.search_time_s) if float(log.search_time_s) >= 0.0 else 0.0
        )
        hud_total = float(cumulative_search_time_s) + trial_search
        search_hud_stim = visual.TextStim(
            win,
            text=format_search_time_hud_text(hud_total),
            pos=(0, sh * 0.40),
            height=max(16.0, int(24 * sc)),
            color=(0.55, 0.55, 0.55),
            units="pix",
            wrapWidth=sw * 0.88,
            font=INSTRUCTIONS_UI_FONT,
            bold=False,
        )
        send_msg(
            tracker,
            "SEARCH_TIME_HUD_TOTAL_S {0:.4f}".format(hud_total),
        )
    choice_clock = core.Clock()
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            log.aborted = True
            log.extra["reason"] = "quit_t_esc"
            _trial_append_event(log, "quit_t_esc", "")
            _trial_append_event(log, "trial_abort", "")
            return log
        keys = kb.getKeys(keyList=["left", "right"], waitRelease=False)
        if keys:
            k = keys[0].name
            log.rt_discrete_s = choice_clock.getTime()
            if k == "left":
                log.discrete_key = "Left"
            else:
                log.discrete_key = "Right"
            _trial_append_event(log, "discrete_choice", k)
            break
        prompt.draw()
        draw_stim_batch(arrow_hint_stims)
        if search_hud_stim is not None:
            search_hud_stim.draw()
        win.flip()

    log.accuracy = (
        "1"
        if log.discrete_key.lower() == (log.correctkey or "").strip().lower()
        else "0"
    )
    send_msg(
        tracker,
        f"DISCRETE_RESPONSE key {log.discrete_key} correct {log.correctkey} acc {log.accuracy}",
    )
    if not show_discrete_response_feedback(win, kb, tracker, log.accuracy == "1"):
        log.aborted = True
        log.extra["reason"] = "quit_t_esc"
        _trial_append_event(log, "quit_t_esc", "")
        _trial_append_event(log, "trial_abort", "")
        return log
    _trial_append_event(
        log,
        "correct_feedback" if log.accuracy == "1" else "incorrect_feedback",
        "",
    )

    _trial_append_event(log, "trial_end", "")
    return log


_SESSION_CSV_FIELDNAMES = (
    "participant_id",
    "experiment",
    "session_condition",
    "session_run",
    "left_panel",
    "right_panel",
    "target_side",
    "trial",
    "trial_id",
    "target_t_x",
    "target_t_y",
    "stim_layout_seed",
    "condition_trial",
    "first_side_choice",
    "forage_time_left_s",
    "forage_time_right_s",
    "cod_switch_count",
    "cod_grey_ms",
    "search_max_s",
    "search_time_s",
    "target_found",
    "search_timed_out",
    "correctkey",
    "discrete_key",
    "accuracy",
    "rt_discrete_s",
    "aborted",
    "reason",
)

_SEGMENT_CSV_FIELDNAMES = (
    "participant_id",
    "experiment",
    "session_condition",
    "session_run",
    "trial",
    "segment_index",
    "side",
    "segment_start_s",
    "segment_end_s",
    "duration_s",
    "switched_to",
    "trial_total_segment_time_s",
    "trial_switch_count",
    "aborted_trial",
    "target_side",
    "found_target_during_segment",
)

_EVENT_CSV_FIELDNAMES = (
    "participant_id",
    "experiment",
    "session_condition",
    "session_run",
    "trial",
    "event_index",
    "event_type",
    "side",
    "t_trial_s",
    "aborted_trial",
)


def enrich_segment_rows_for_trial(
    log: TrialLog,
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    trial_num: int,
) -> List[Dict[str, Any]]:
    total = sum(float(s.get("duration_s", 0) or 0) for s in log.segment_rows)
    total_r = round(total, 6)
    aborted_v = 1 if log.aborted else 0
    out: List[Dict[str, Any]] = []
    for s in log.segment_rows:
        out.append(
            {
                "participant_id": participant_safe,
                "experiment": experiment_number,
                "session_condition": session_condition,
                "session_run": session_run,
                "trial": trial_num,
                "segment_index": s["segment_index"],
                "side": s["side"],
                "segment_start_s": s["segment_start_s"],
                "segment_end_s": s["segment_end_s"],
                "duration_s": s["duration_s"],
                "switched_to": s["switched_to"],
                "trial_total_segment_time_s": total_r,
                "trial_switch_count": log.cod_switch_count,
                "aborted_trial": aborted_v,
                "target_side": log.target_side,
                "found_target_during_segment": s["found_target_during_segment"],
            }
        )
    return out


def enrich_event_rows_for_trial(
    log: TrialLog,
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    trial_num: int,
) -> List[Dict[str, Any]]:
    aborted_v = 1 if log.aborted else 0
    return [
        {
            "participant_id": participant_safe,
            "experiment": experiment_number,
            "session_condition": session_condition,
            "session_run": session_run,
            "trial": trial_num,
            "event_index": e["event_index"],
            "event_type": e["event_type"],
            "side": e["side"],
            "t_trial_s": e["t_trial_s"],
            "aborted_trial": aborted_v,
        }
        for e in log.event_rows
    ]


def _write_session_csv(rows_out: List[Dict[str, Any]], csv_path: str) -> None:
    """Grava CSV de sessão; `rows_out` vazio gera só cabeçalho (T+Esc sem trials)."""
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    fieldnames = list(_SESSION_CSV_FIELDNAMES)
    for row in rows_out:
        for k in row:
            if k not in fieldnames:
                fieldnames.append(k)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=fieldnames, extrasaction="ignore", restval=""
        )
        w.writeheader()
        for row in rows_out:
            w.writerow(row)


def _write_fixed_csv(
    rows_out: List[Dict[str, Any]],
    csv_path: str,
    fieldnames: Tuple[str, ...],
) -> None:
    os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=list(fieldnames), extrasaction="ignore", restval=""
        )
        w.writeheader()
        for row in rows_out:
            w.writerow(row)


def _tsv_block_lines(rows_out: List[Dict[str, Any]], title: str) -> List[str]:
    lines = [title, ""]
    if not rows_out:
        lines.append("(sem linhas)")
        return lines
    keys: List[str] = []
    for row in rows_out:
        for k in row:
            if k not in keys:
                keys.append(k)
    lines.append("\t".join(keys))
    for row in rows_out:
        lines.append("\t".join(str(row.get(k, "")) for k in keys))
    return lines


def _write_session_txt(
    rows_out: List[Dict[str, Any]],
    segment_rows: List[Dict[str, Any]],
    event_rows: List[Dict[str, Any]],
    txt_path: str,
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    ts: str,
    session_note: str,
    dummy_mode: bool,
) -> None:
    """TXT: cabeçalho + três blocos (trials / segmentos / eventos) em TSV."""
    os.makedirs(os.path.dirname(txt_path) or ".", exist_ok=True)
    mode = "dummy (rato)" if dummy_mode else "EyeLink"
    lines = [
        "Forrageamento visual — Experimento {0}".format(int(experiment_number)),
        "Código do participante (ficheiros): {0}".format(participant_safe),
        "Experimento (1 ou 2): {0}".format(int(experiment_number)),
        "Sessão (condição 1–5): {0}".format(session_condition),
        "Repetição desta condição: {0}".format(session_run),
        "Data/hora gravação: {0}".format(ts),
        "Distância olhos–ecrã (protocolo): {0} cm".format(int(VIEWING_DISTANCE_CM)),
        "Modo: {0}".format(mode),
        "Nota: {0}".format(session_note or "—"),
        "",
    ]
    lines.extend(_tsv_block_lines(rows_out, "--- Nível 1: trials (TSV) ---"))
    lines.append("")
    lines.extend(_tsv_block_lines(segment_rows, "--- Nível 2: segmentos (TSV) ---"))
    lines.append("")
    lines.extend(_tsv_block_lines(event_rows, "--- Nível 3: eventos (TSV) ---"))
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


def _xlsx_write_meta_and_table(
    ws: Any,
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    ts: str,
    session_note: str,
    dummy_mode: bool,
    rows_out: List[Dict[str, Any]],
    base_fieldnames: Optional[Tuple[str, ...]] = None,
) -> None:
    mode = "dummy" if dummy_mode else "eyelink"
    ws.append(["participant_id", participant_safe])
    ws.append(["experiment", int(experiment_number)])
    ws.append(["session_condition", session_condition])
    ws.append(["session_run", session_run])
    ws.append(["timestamp", ts])
    ws.append(["distancia_olhos_ecran_cm_protocolo", int(VIEWING_DISTANCE_CM)])
    ws.append(["modo", mode])
    ws.append(["nota", session_note or ""])
    ws.append([])
    if rows_out:
        if base_fieldnames is not None:
            keys: List[str] = list(base_fieldnames)
            for row in rows_out:
                for k in row:
                    if k not in keys:
                        keys.append(k)
        else:
            keys = []
            for row in rows_out:
                for k in row:
                    if k not in keys:
                        keys.append(k)
        ws.append(keys)
        for row in rows_out:
            ws.append([row.get(k, "") for k in keys])


def _write_session_xlsx(
    rows_out: List[Dict[str, Any]],
    segment_rows: List[Dict[str, Any]],
    event_rows: List[Dict[str, Any]],
    xlsx_path: str,
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    ts: str,
    session_note: str,
    dummy_mode: bool,
) -> bool:
    try:
        from openpyxl import Workbook
    except Exception:
        return False
    os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
    wb = Workbook()
    ws_t = wb.active
    ws_t.title = "trials"
    _xlsx_write_meta_and_table(
        ws_t,
        participant_safe,
        experiment_number,
        session_condition,
        session_run,
        ts,
        session_note,
        dummy_mode,
        rows_out,
        base_fieldnames=_SESSION_CSV_FIELDNAMES,
    )
    ws_s = wb.create_sheet("segments")
    _xlsx_write_meta_and_table(
        ws_s,
        participant_safe,
        experiment_number,
        session_condition,
        session_run,
        ts,
        session_note,
        dummy_mode,
        segment_rows,
        base_fieldnames=_SEGMENT_CSV_FIELDNAMES,
    )
    ws_e = wb.create_sheet("events")
    _xlsx_write_meta_and_table(
        ws_e,
        participant_safe,
        experiment_number,
        session_condition,
        session_run,
        ts,
        session_note,
        dummy_mode,
        event_rows,
        base_fieldnames=_EVENT_CSV_FIELDNAMES,
    )
    wb.save(xlsx_path)
    return True


def save_session_exports(
    rows_out: List[Dict[str, Any]],
    segment_rows: List[Dict[str, Any]],
    event_rows: List[Dict[str, Any]],
    participant_safe: str,
    experiment_number: int,
    session_condition: int,
    session_run: int,
    ts: str,
    session_note: str,
    dummy_mode: bool,
) -> Tuple[str, str, str, str, Optional[str]]:
    """
    Grava três níveis: trials (.csv), segmentos (_segments.csv), eventos (_events.csv);
    TXT com três blocos; XLSX com folhas trials / segments / events (se openpyxl).
    Devolve (csv_trials, csv_segments, csv_events, txt_path, xlsx_path|None).
    """
    os.makedirs(_DATA_FOLDER, exist_ok=True)
    base = os.path.join(
        _DATA_FOLDER,
        "forrageamento_exp{0}_{1}_s{2}_r{3}_{4}".format(
            int(experiment_number),
            participant_safe,
            session_condition,
            session_run,
            ts,
        ),
    )
    csv_path = base + ".csv"
    csv_seg = base + "_segments.csv"
    csv_evt = base + "_events.csv"
    txt_path = base + ".txt"
    xlsx_path = base + ".xlsx"
    _write_session_csv(rows_out, csv_path)
    _write_fixed_csv(segment_rows, csv_seg, _SEGMENT_CSV_FIELDNAMES)
    _write_fixed_csv(event_rows, csv_evt, _EVENT_CSV_FIELDNAMES)
    _write_session_txt(
        rows_out,
        segment_rows,
        event_rows,
        txt_path,
        participant_safe,
        experiment_number,
        session_condition,
        session_run,
        ts,
        session_note,
        dummy_mode,
    )
    xlsx_ok = _write_session_xlsx(
        rows_out,
        segment_rows,
        event_rows,
        xlsx_path,
        participant_safe,
        experiment_number,
        session_condition,
        session_run,
        ts,
        session_note,
        dummy_mode,
    )
    try:
        import forrageamento_analysis_export as _fa

        p_at, p_as, p_ag = _fa.write_exp1_analysis_exports(
            base, rows_out, segment_rows
        )
        print(
            "  CSV análise (trials): {0}\n  CSV análise (sessão): {1}\n"
            "  CSV análise (segmentos): {2}".format(p_at, p_as, p_ag)
        )
    except Exception as _ex:
        print(
            "  Aviso: exportação _analysis_* omitida ({0})".format(_ex),
            file=sys.stderr,
        )
    print(
        "Dados gravados:\n  CSV trials: {0}\n  CSV segmentos: {1}\n"
        "  CSV eventos: {2}\n  TXT: {3}".format(csv_path, csv_seg, csv_evt, txt_path)
    )
    if xlsx_ok:
        print("  XLSX: {0} (folhas: trials, segments, events)".format(xlsx_path))
    else:
        print(
            "  XLSX: omitido (instale: pip install openpyxl)",
            file=sys.stderr,
        )
    return csv_path, csv_seg, csv_evt, txt_path, xlsx_path if xlsx_ok else None


def run_instructions(win: visual.Window, kb, tracker: Any):
    """Texto branco (PT-BR); caixa de exemplo com L e T verde gerada no código."""
    sw, sh = float(win.size[0]), float(win.size[1])
    h_title, h_body, h_footer = instruction_font_heights(sh)
    w_wrap = min(sw * 0.86, 1040.0)
    # Coluna compacta: blocos mais próximos (pos em frações de sh, origem no centro)
    y_title = sh * 0.34
    y_intro = sh * 0.24
    y_task = sh * 0.13
    y_example = sh * -0.05
    y_rules = sh * -0.26
    y_footer = sh * -0.38
    ex_w, ex_h = instruction_example_panel_size(sw, sh, frac=0.26)

    instr_title = make_instruction_text(
        win,
        "LEIA AS INSTRUÇÕES COM ATENÇÃO!",
        y_title,
        h_title,
        w_wrap,
    )
    instr_intro = make_instruction_text(
        win,
        (
            "Sua tarefa é encontrar a letra T entre várias letras L e indicar em qual lado ela estava.\n"
            "1. PREPARAÇÃO: Olhe para o alvo central até ele desaparecer. Depois, olhe para\n"
            "a esquerda ou para a direita para escolher por onde começar."
        ),
        y_intro,
        h_body,
        w_wrap,
    )
    instr_task = make_instruction_text(
        win,
        (
            "2. BUSCA: Apenas um lado fica visível por vez. Para trocar, olhe para o lado vazio e aguarde.\n"
            "Procure o T entre os L (exemplo abaixo). O verde é só uma ajuda nesta ilustração:\n"
            "durante a tarefa, as letras aparecerão todas em tons de cinza."
        ),
        y_task,
        h_body,
        w_wrap,
    )
    example_stims = build_instruction_example_panel_stims(
        win, 0.0, y_example, ex_w, ex_h, font=INSTRUCTIONS_UI_FONT
    )
    instr_rules = make_instruction_text(
        win,
        (
            "3. RESPOSTA: Ao encontrar o T, mantenha o olhar sobre ele e pressione ESPAÇO.\n"
            "O pequeno ponto acompanha a posição do seu olhar e ajuda a centralizá-lo no T. "
            "Há um tempo limite para encontrar o T; se não o achar a tempo, a tarefa avançará.\n"
            "Depois, indique o lado em que ele estava:  ←  ESQUERDA    |    DIREITA  →\n"
            "A tela ficará verde se você acertar e cinza se errar.\n"
            "Tente acertar o máximo possível."
        ),
        y_rules,
        h_body,
        w_wrap,
    )
    instr_footer = make_instruction_text(
        win,
        "PRESSIONE A BARRA DE ESPAÇO PARA COMEÇAR.",
        y_footer,
        h_footer,
        w_wrap,
    )
    send_msg(tracker, "PHASE instructions")
    flush_keyboard(kb)
    core.wait(0.15)
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            return False
        keys = kb.getKeys(keyList=["space"], waitRelease=False)
        if keys:
            break
        instr_title.draw()
        instr_intro.draw()
        instr_task.draw()
        draw_stim_batch(example_stims)
        instr_rules.draw()
        instr_footer.draw()
        win.flip()
    flush_keyboard(kb)
    return True


def run_instructions_exp2(win: visual.Window, kb, tracker: Any):
    """Instruções Exp2: pontos, decaimento, sem fase de setas após cada alvo."""
    sw, sh = float(win.size[0]), float(win.size[1])
    h_title, h_body, h_footer = instruction_font_heights(sh)
    w_wrap = min(sw * 0.84, 1000.0)

    instr_title = make_instruction_text(
        win, "LEIA AS INSTRUÇÕES COM ATENÇÃO!", sh * 0.36, h_title, w_wrap
    )
    pre_lines = (
        "Sua tarefa é ganhar pontos encontrando a letra T entre várias letras L.\n"
        "1. BUSCA: Há duas áreas, uma à esquerda e outra à direita, mas apenas uma fica visível por vez.\n"
        "Para trocar de área, olhe para o lado vazio e aguarde. Em cada área existe um T."
    )
    instr_pre = make_instruction_text(win, pre_lines, sh * 0.22, h_body, w_wrap)
    ex_w, ex_h = instruction_example_panel_size(sw, sh, frac=0.28)
    example_stims = build_instruction_example_panel_stims(
        win, 0.0, sh * -0.02, ex_w, ex_h, font=INSTRUCTIONS_UI_FONT
    )
    post_lines = (
        "O T verde é só uma ajuda nesta ilustração; durante a tarefa, todas as letras terão tons de cinza.\n"
        "O pequeno ponto acompanha a posição do seu olhar e ajuda a centralizá-lo no T.\n"
        "2. PONTUAR: Ao encontrar o T, mantenha o olhar sobre ele e pressione ESPAÇO. Cada acerto vale 1 ponto.\n"
        "3. ATENÇÃO: As letras L desaparecem aos poucos. Após um ponto, somente a área em que você\n"
        "pontuou é reiniciada; a outra continua como estava.\n"
        "Tente ganhar o maior número de pontos possível."
    )
    instr_post = make_instruction_text(win, post_lines, sh * -0.24, h_body, w_wrap)
    instr_footer = make_instruction_text(
        win,
        "PRESSIONE A BARRA DE ESPAÇO PARA INICIAR A SESSÃO.",
        sh * -0.38,
        h_footer,
        w_wrap,
    )
    send_msg(tracker, "PHASE instructions_exp2")
    flush_keyboard(kb)
    core.wait(0.15)
    while True:
        poll_global_quit_t_esc(kb, tracker, win)
        if QUIT_SAVE_REQUESTED:
            return False
        keys = kb.getKeys(keyList=["space"], waitRelease=False)
        if keys:
            break
        instr_title.draw()
        instr_pre.draw()
        draw_stim_batch(example_stims)
        instr_post.draw()
        instr_footer.draw()
        win.flip()
    flush_keyboard(kb)
    return True


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dummy", action="store_true", help="Use mouse as gaze; skip startup menu")
    parser.add_argument(
        "--skip-mode-menu",
        action="store_true",
        help="No E/M menu: use EyeLink unless --dummy / FORR_DUMMY / RUN_DUMMY_WHEN_NO_ARGS",
    )
    parser.add_argument(
        "--eyelink-host",
        default=DEFAULT_EYELINK_HOST,
        help="IP do EyeLink host (por defeito: %s ou FORR_EYELINK_HOST)" % DEFAULT_EYELINK_HOST,
    )
    parser.add_argument("--trials", type=int, default=120, help="Number of trials to run")
    parser.add_argument(
        "--stim-dir",
        default="",
        help="Ignorado: estímulos são L/T em cinza gerados no código (compatibilidade de linha de comando).",
    )
    parser.add_argument(
        "--conditions",
        default=None,
        help="CSV de condições (por defeito: conditions_exp1.csv ou conditions_exp2.csv conforme --experiment)",
    )
    parser.add_argument(
        "--experiment",
        type=int,
        default=None,
        choices=[1, 2],
        metavar="1|2",
        help="Experimento 1 ou 2 (FORR_EXPERIMENT); com --participant completo evita diálogo",
    )
    parser.add_argument(
        "--no-shuffle-trials",
        action="store_true",
        help="Manter a ordem das linhas do CSV (por defeito a ordem é permutada por sessão)",
    )
    parser.add_argument(
        "--window-width",
        type=int,
        default=None,
        help="Largura lógica da janela (FORR_WIN_W ou %d)" % DEFAULT_WINDOW_W,
    )
    parser.add_argument(
        "--window-height",
        type=int,
        default=None,
        help="Altura lógica (FORR_WIN_H ou %d)" % DEFAULT_WINDOW_H,
    )
    parser.add_argument(
        "--resolution",
        default=None,
        metavar="WxH",
        help=(
            "Par largura×altura (ex.: 1680x1050; predef.: %dx%d). Alternativa: FORR_RESOLUTION."
            % (DEFAULT_WINDOW_W, DEFAULT_WINDOW_H)
        ),
    )
    parser.add_argument(
        "--participant",
        default=None,
        help="Código do participante (com --session e --session-run evita diálogo)",
    )
    parser.add_argument(
        "--session",
        type=int,
        default=None,
        metavar="N",
        help="Condição/sessão: Exp1 → 1–5; Exp2 → 1–10 (FORR_SESSION)",
    )
    parser.add_argument(
        "--session-run",
        type=int,
        default=None,
        metavar="N",
        help="Nª vez que esta condição é corrida (1, 2, 3…); FORR_SESSION_RUN",
    )
    parser.add_argument(
        "--search-max-s",
        type=float,
        default=None,
        metavar="SEC",
        help="Exp1: tempo máx. (s) por trial na procura do T, do olhar centrado até à "
        "pergunta lateral (FORR_SEARCH_MAX_S / forrageamento_prefs.json; predef. %.0f)"
        % DEFAULT_SEARCH_MAX_S,
    )
    parser.add_argument(
        "--cod-grey-ms",
        type=float,
        default=None,
        metavar="MS",
        help=(
            'Change over delay / pausa COD (ms) ao trocar de alternativa '
            '(FORR_COD_GREY_MS; predef. %.0f = "Padrão")'
        )
        % DEFAULT_COD_GREY_MS,
    )
    parser.add_argument(
        "--no-gaze-dot",
        action="store_true",
        default=False,
        help="Não mostrar o ponto do olhar (padrão: mostrar; FORR_SHOW_GAZE_DOT=0)",
    )
    parser.add_argument(
        "--show-search-time-hud",
        action="store_true",
        default=False,
        help=(
            "Exp1: mostrar na ecrã ←/→ a soma do tempo de procura dos trials "
            "(padrão: off; FORR_SHOW_SEARCH_TIME_HUD=1)"
        ),
    )
    parser.add_argument(
        "--exp2-duration-s",
        type=float,
        default=None,
        metavar="SEC",
        help=(
            "Exp2: tempo total (s) após instruções "
            '(FORR_EXP2_DURATION_S / prefs; predef. %.0f = "Padrão")'
        )
        % DEFAULT_EXP2_DURATION_S,
    )
    parser.add_argument(
        "--n-l-left",
        type=int,
        default=None,
        metavar="N",
        help="Exp2: número inicial de L no painel esquerdo (omissão = CSV da condição)",
    )
    parser.add_argument(
        "--n-l-right",
        type=int,
        default=None,
        metavar="N",
        help="Exp2: número inicial de L no painel direito (omissão = CSV da condição)",
    )
    args = parser.parse_args()
    reset_session_quit_flags()

    meta = resolve_experiment_metadata(args)
    if meta is None:
        print("Sessão cancelada ou metadados incompletos.")
        return 0
    (
        participant_raw,
        session_condition,
        session_run,
        experiment_number,
        gui_win_w,
        gui_win_h,
        search_max_s,
        cod_grey_ms,
        show_gaze_dot,
        show_search_time_hud,
        exp2_duration_s,
        n_L_left,
        n_L_right,
    ) = meta
    participant_safe = sanitize_participant_id(participant_raw)

    conditions_path = resolve_conditions_path(args, experiment_number)
    if not os.path.isfile(conditions_path):
        if experiment_number == 2 and args.conditions is None:
            print(
                "Aviso: {0} não encontrado; a usar conditions_exp1.csv.".format(
                    _CONDITIONS_FILE_EXP2_SESSIONS
                ),
                file=sys.stderr,
            )
            conditions_path = _CONDITIONS_FILE_EXP1
        if not os.path.isfile(conditions_path):
            print(
                "Ficheiro de condições em falta: {0}".format(conditions_path),
                file=sys.stderr,
            )
            return 1

    forced_dummy = args.dummy or os.environ.get("FORR_DUMMY", "") == "1"
    legacy_dummy_no_args = (
        args.skip_mode_menu
        and not args.dummy
        and RUN_DUMMY_WHEN_NO_ARGS
        and len(sys.argv) <= 1
    )
    dummy_mode = forced_dummy or legacy_dummy_no_args

    with open(conditions_path, newline="", encoding="utf-8") as f:
        trial_list_full = list(csv.DictReader(f))
    if experiment_number == 2:
        trial_list = list(trial_list_full)
    elif args.no_shuffle_trials:
        trial_list = list(trial_list_full)
    else:
        order_rng = random.Random(
            trial_order_seed(
                participant_safe, experiment_number, session_condition, session_run
            )
        )
        order_ix = list(range(len(trial_list_full)))
        order_rng.shuffle(order_ix)
        trial_list = [trial_list_full[i] for i in order_ix]

    win_w, win_h = resolve_window_size(args, gui_win_w, gui_win_h)
    win = visual.Window(
        size=(win_w, win_h),
        fullscr=True,
        color=BACKGROUND_COLOR,
        units="pix",
    )
    kb = make_keyboard()
    mouse = event.Mouse(win=win)
    flush_keyboard(kb)

    tracker = None
    edf_name = make_valid_edf_name(participant_safe)
    print(
        "Participante: {0}  |  Experimento: {1}  |  Sessão: {2}  |  Repetição: {3}  |  EDF: {4}  |  Janela: {5}×{6}".format(
            participant_safe,
            experiment_number,
            session_condition,
            session_run,
            edf_name,
            win_w,
            win_h,
        )
    )
    if experiment_number != 2 and not args.no_shuffle_trials:
        print(
            "Ordem dos trials: permutada (reprodutível para este participante/sessão/repetição)."
        )

    if not forced_dummy and not args.skip_mode_menu:
        choice = run_mode_selection(win, kb, args.eyelink_host)
        if choice == "quit_save":
            run_experiment_end_screen(win, kb, tracker)
            ts = time.strftime("%Y%m%d_%H%M%S")
            save_session_exports(
                [],
                [],
                [],
                participant_safe,
                experiment_number,
                session_condition,
                session_run,
                ts,
                "T+Esc no menu (sem trials)",
                dummy_mode,
            )
            print("Sessão terminada (T+Esc); dados gravados após Espaço no ecrã final.")
            shutdown_eyelink(tracker, edf_name, _DATA_FOLDER)
            win.close()
            core.quit()
            return 0
        if choice is None:
            print("Sessão cancelada (ESC no menu).")
            win.close()
            core.quit()
            return 0
        dummy_mode = choice

    if not dummy_mode:
        try:
            tracker = setup_eyelink(win, edf_name, args.eyelink_host)
            send_msg(tracker, "PARTICIPANT_ID {0}".format(participant_safe))
            send_msg(tracker, "EXPERIMENT {0}".format(int(experiment_number)))
            send_msg(tracker, "SESSION_CONDITION {0}".format(session_condition))
            send_msg(tracker, "SESSION_RUN {0}".format(session_run))
            if experiment_number == 1:
                send_msg(tracker, "SEARCH_MAX_S {0:.4f}".format(float(search_max_s)))
            send_msg(tracker, "COD_GREY_MS {0:.1f}".format(float(cod_grey_ms)))
            send_msg(
                tracker,
                "SHOW_GAZE_DOT {0}".format(1 if show_gaze_dot else 0),
            )
        except Exception as e:
            err_txt = str(e)
            print(
                "EyeLink setup failed ({0}); para testar sem hardware use --dummy".format(err_txt),
                file=sys.stderr,
            )
            show_error_screen(
                win,
                kb,
                "Falha na ligação / calibração EyeLink",
                (
                    "Detalhe: {0}\n\n"
                    "IP tentado: {1}\n\n"
                    "Opções:\n"
                    "• Volte a correr o programa e no primeiro ecrã escolha M (rato).\n"
                    "• Ou: linha de comando com --dummy\n"
                    "• IP errado: --eyelink-host IP ou FORR_EYELINK_HOST\n"
                    "• Automação sem menu: --skip-mode-menu (e opcionalmente --dummy)"
                ).format(err_txt, args.eyelink_host),
            )
            win.close()
            core.quit()
            return 1
    else:
        print(
            "DUMMY MODE: mouse position simulates gaze. Participante: {0}".format(
                participant_safe
            )
        )

    try:
        if experiment_number == 2:
            import forrageamento_exp2_session as fe2

            session_row = None
            for r in trial_list_full:
                try:
                    if int((r.get("session") or "0").strip()) == int(session_condition):
                        session_row = r
                        break
                except ValueError:
                    continue
            if session_row is None and trial_list_full:
                session_row = trial_list_full[0]
            if session_row is None:
                print("CSV do Experimento 2 sem linhas válidas.", file=sys.stderr)
                shutdown_eyelink(tracker, edf_name, _DATA_FOLDER)
                win.close()
                core.quit()
                return 1
            print(
                "Experimento 2 — sessão contínua | condição CSV: {0} | "
                "COD: {1:.0f} ms | duração: {2:.0f} s | L: {3}/{4}".format(
                    (session_row.get("ratio_label") or "").strip() or "?",
                    float(cod_grey_ms),
                    float(exp2_duration_s),
                    n_L_left if n_L_left is not None else session_row.get("n_L_left", "?"),
                    n_L_right if n_L_right is not None else session_row.get("n_L_right", "?"),
                )
            )
            fe2.run_exp2_session_wrapper(
                win,
                kb,
                mouse,
                tracker,
                dummy_mode,
                participant_safe,
                session_condition,
                session_run,
                session_row,
                cod_grey_ms=float(cod_grey_ms),
                show_gaze_dot=bool(show_gaze_dot),
                duration_s=float(exp2_duration_s),
                n_L_left=n_L_left,
                n_L_right=n_L_right,
            )
            shutdown_eyelink(tracker, edf_name, _DATA_FOLDER)
            win.close()
            core.quit()
            return 0

        ok = run_instructions(win, kb, tracker)
        if not ok:
            # T+Esc nas instruções (sem trials completos)
            run_experiment_end_screen(win, kb, tracker)
            ts = time.strftime("%Y%m%d_%H%M%S")
            save_session_exports(
                [],
                [],
                [],
                participant_safe,
                experiment_number,
                session_condition,
                session_run,
                ts,
                "T+Esc nas instruções (sem trials)",
                dummy_mode,
            )
            print("Sessão terminada (T+Esc); dados gravados após Espaço no ecrã final.")
            shutdown_eyelink(tracker, edf_name, _DATA_FOLDER)
            win.close()
            core.quit()
            return 0

        rows_out: List[Dict[str, Any]] = []
        segment_rows_all: List[Dict[str, Any]] = []
        event_rows_all: List[Dict[str, Any]] = []
        n_run = min(args.trials, len(trial_list))
        ended_by_t_esc = False
        print(
            "Experimento 1 — {0} trials, sem limite de tempo total | tempo máx. de procura "
            "do T por trial: {1:.1f} s (desde o olhar centrado até à pergunta lateral) | "
            "COD: {2:.0f} ms | HUD tempo total procura: {3}".format(
                n_run,
                float(search_max_s),
                float(cod_grey_ms),
                "ligado" if show_search_time_hud else "desligado",
            )
        )

        cumulative_search_time_s = 0.0
        for t_ix in range(n_run):
            trial = trial_list[t_ix]
            stim_seed = stimulus_layout_seed(
                participant_safe,
                experiment_number,
                session_condition,
                session_run,
                t_ix + 1,
                trial.get("trial", ""),
            )
            print(
                "--- Trial {0}/{1} (condição CSV trial={2}) ---".format(
                    t_ix + 1,
                    n_run,
                    (trial.get("trial") or "").strip() or "?",
                )
            )
            log = run_trial(
                win,
                kb,
                mouse,
                tracker,
                trial,
                t_ix + 1,
                dummy_mode,
                stim_seed,
                search_max_s=float(search_max_s),
                cod_grey_ms=float(cod_grey_ms),
                show_gaze_dot=bool(show_gaze_dot),
                show_search_time_hud=bool(show_search_time_hud),
                cumulative_search_time_s=float(cumulative_search_time_s),
            )
            if float(log.search_time_s) >= 0.0:
                cumulative_search_time_s += float(log.search_time_s)
            tn = t_ix + 1
            segment_rows_all.extend(
                enrich_segment_rows_for_trial(
                    log,
                    participant_safe,
                    experiment_number,
                    session_condition,
                    session_run,
                    tn,
                )
            )
            event_rows_all.extend(
                enrich_event_rows_for_trial(
                    log,
                    participant_safe,
                    experiment_number,
                    session_condition,
                    session_run,
                    tn,
                )
            )
            lp, rp = panels_lt_from_target_side(log.target_side)
            row = {
                "participant_id": participant_safe,
                "experiment": experiment_number,
                "session_condition": session_condition,
                "session_run": session_run,
                "left_panel": lp,
                "right_panel": rp,
                "target_side": log.target_side,
                "trial": t_ix + 1,
                "trial_id": log.trial_id,
                "target_t_x": ""
                if log.target_t_x is None
                else str(log.target_t_x),
                "target_t_y": ""
                if log.target_t_y is None
                else str(log.target_t_y),
                "stim_layout_seed": stim_seed,
                "condition_trial": trial.get("trial", ""),
                "first_side_choice": log.active_side_after_choice,
                "forage_time_left_s": log.forage_time_left_s,
                "forage_time_right_s": log.forage_time_right_s,
                "cod_switch_count": log.cod_switch_count,
                "cod_grey_ms": round(float(cod_grey_ms), 1),
                "search_max_s": round(float(search_max_s), 4),
                "search_time_s": log.search_time_s,
                "target_found": "1" if log.target_found else "0",
                "search_timed_out": "1" if log.search_timed_out else "0",
                "correctkey": log.correctkey,
                "discrete_key": log.discrete_key,
                "accuracy": log.accuracy,
                "rt_discrete_s": log.rt_discrete_s,
                "aborted": log.aborted,
            }
            row.update(log.extra)
            rows_out.append(row)
            if log.extra.get("reason") == "quit_t_esc" or QUIT_SAVE_REQUESTED:
                print("Sessão terminada (T+Esc).")
                ended_by_t_esc = True
                break

        run_experiment_end_screen(win, kb, tracker)
        ts = time.strftime("%Y%m%d_%H%M%S")
        note = (
            "T+Esc durante trials (dados parciais)"
            if ended_by_t_esc
            else "Todos os trials da sessão concluídos"
        )
        save_session_exports(
            rows_out,
            segment_rows_all,
            event_rows_all,
            participant_safe,
            experiment_number,
            session_condition,
            session_run,
            ts,
            note,
            dummy_mode,
        )
        print("Dados gravados após Espaço no ecrã final.")

    except Exception as run_exc:
        import traceback

        tb = traceback.format_exc()
        print(tb, file=sys.stderr)
        show_error_screen(
            win,
            kb,
            "Erro durante o experimento",
            tb[-500:] if len(tb) > 500 else tb,
        )

    shutdown_eyelink(tracker, edf_name, _DATA_FOLDER)
    win.close()
    core.quit()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
